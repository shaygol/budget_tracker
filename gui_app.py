"""
PyQt5 GUI application for Budget Tracker.
Provides a modern interface for transaction processing with bilingual support.
"""

import sys
import os
import shutil
from pathlib import Path
from datetime import datetime
from typing import List, Optional, Callable
import pandas as pd
import logging

# Suppress matplotlib debug logging
logging.getLogger('matplotlib').setLevel(logging.WARNING)
logging.getLogger('PIL').setLevel(logging.WARNING)

# Fix for PyQt5 platform plugin issue
import PyQt5
plugin_path = os.path.join(os.path.dirname(PyQt5.__file__), 'Qt5', 'plugins', 'platforms')
os.environ['QT_QPA_PLATFORM_PLUGIN_PATH'] = plugin_path  # noqa: E402

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QListWidget, QTabWidget, QTableWidget,
    QTableWidgetItem, QProgressBar, QFileDialog, QMessageBox,
    QDialog, QComboBox, QDialogButtonBox, QTextEdit, QHeaderView, QLineEdit
)  # noqa: E402
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QTimer, QObject  # noqa: E402
from PyQt5.QtGui import QFont, QDragEnterEvent, QDropEvent, QColor, QIcon  # noqa: E402

from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas  # noqa: E402
from matplotlib.figure import Figure  # noqa: E402

from src.config import TRANSACTIONS_DIR, CATEGORIES_FILE_PATH, DASHBOARD_FILE_PATH, OUTPUT_DIR, LOG_FILE_NAME, ARCHIVE_DIR  # noqa: E402
from src.file_manager import ensure_dirs, load_transaction_files  # noqa: E402
from src.logger import setup_logging  # noqa: E402
from src.normalizer import Normalizer  # noqa: E402
from src.category_manager import CategoryManager  # noqa: E402
from src.previewer import Previewer  # noqa: E402
from src.dashboard_writer import DashboardWriter  # noqa: E402
from src.config import get_log_level  # noqa: E402
from src.translations import Translations  # noqa: E402
from src.file_utils import is_file_locked, validate_dashboard_integrity, get_user_friendly_error, check_file_permissions  # noqa: E402


class ProcessThread(QThread):
    """Background thread for processing transactions."""
    progress = pyqtSignal(str)
    finished = pyqtSignal(pd.DataFrame, bool)
    error = pyqtSignal(str)
    category_needed = pyqtSignal(str, list, dict, str)  # merchant, choices, sample_data, progress_text
    log_message = pyqtSignal(str, str)  # level, message

    def __init__(self, translations):
        """
        Initialize the processing thread.

        Args:
            translations: Translations object for localized messages
        """
        super().__init__()
        self.translations = translations
        self.category_response = None
        self.response_ready = False
        self._should_stop = False

        # Timeout handling
        self.timeout_timer = QTimer()
        self.timeout_timer.setSingleShot(True)
        self.timeout_timer.timeout.connect(self.handle_timeout)
        self.started.connect(self.start_timeout)
        self.finished.connect(self.stop_timeout)

    def run(self):
        """
        Main thread execution method.

        Loads transaction files, normalizes data, maps categories (with GUI interaction
        for unknown merchants), generates preview, and emits finished signal with results.
        """
        import traceback
        try:
            if self._should_stop:
                return

            self.log_message.emit('INFO', '=== Starting transaction processing ===')
            self.progress.emit(self.translations.get('processing'))

            # Load files
            self.log_message.emit('INFO', 'Loading transaction files...')
            dfs = load_transaction_files(TRANSACTIONS_DIR)
            if not dfs:
                self.error.emit(self.translations.get('no_files'))
                return

            self.log_message.emit('INFO', f'Loaded {len(dfs)} transaction file(s)')

            # Normalize
            self.log_message.emit('INFO', 'Normalizing transactions...')
            normalizer = Normalizer()
            df = pd.concat(dfs, ignore_index=True)
            df = normalizer.normalize(df)
            self.log_message.emit('INFO', f'Normalized {len(df)} transactions')

            if self._should_stop:
                return

            # Category mapping
            self.log_message.emit('INFO', 'Mapping categories...')
            cat_mgr = CategoryManager(CATEGORIES_FILE_PATH, DASHBOARD_FILE_PATH)
            df = self._map_categories_gui(df, cat_mgr)

            if self._should_stop:
                return

            # Preview
            self.log_message.emit('INFO', 'Generating preview...')
            preview = Previewer()
            summary = preview.preview(df, confirm=False)

            self.log_message.emit('INFO', 'Processing complete!')
            self.finished.emit(summary, True)

        except FileNotFoundError as e:
            error_msg = get_user_friendly_error(e)
            self.log_message.emit('ERROR', f"File not found: {str(e)}")
            self.error.emit(error_msg)
        except PermissionError as e:
            error_msg = get_user_friendly_error(e)
            self.log_message.emit('ERROR', f"Permission error: {str(e)}")
            self.error.emit(error_msg)
        except (IOError, OSError) as e:
            error_msg = get_user_friendly_error(e)
            self.log_message.emit('ERROR', f"IO error: {str(e)}")
            self.error.emit(error_msg)
        except ValueError as e:
            error_msg = get_user_friendly_error(e)
            self.log_message.emit('ERROR', f"Data error: {str(e)}")
            self.error.emit(error_msg)
        except Exception as e:
            error_details = f"{str(e)}\\n\\n{traceback.format_exc()}"
            error_msg = get_user_friendly_error(e)
            self.log_message.emit('ERROR', error_details)
            self.error.emit(error_msg)

    def start_timeout(self):
        """Start the processing timeout timer."""
        from src.config import PROCESSING_TIMEOUT_SECONDS
        self.timeout_timer.start(PROCESSING_TIMEOUT_SECONDS * 1000)

    def stop_timeout(self):
        """Stop the processing timeout timer."""
        self.timeout_timer.stop()

    def handle_timeout(self):
        """Handle timeout event when processing takes too long."""
        self.log_message.emit('ERROR', 'Processing timeout - operation took too long')
        timeout_msg = self.translations.get('timeout_error') or 'Processing timed out'
        self.error.emit(timeout_msg)
        self.stop()

    def stop(self):
        """Request thread to stop."""
        self._should_stop = True

    def _map_categories_gui(self, df: pd.DataFrame, cat_mgr: CategoryManager) -> pd.DataFrame:
        """
        Map categories with GUI interaction for unknown merchants.

        For each merchant not in the category map, prompts the user via GUI dialog
        to select a category/subcategory. Updates the dataframe and saves mappings.

        Args:
            df: DataFrame with transaction data
            cat_mgr: CategoryManager instance for category operations

        Returns:
            DataFrame with category and subcat columns populated
        """
        df = df.copy()

        df['category'] = df['merchant'].map(lambda m: cat_mgr.category_map.get(m, [None, None])[0])
        df['subcat'] = df['merchant'].map(lambda m: cat_mgr.category_map.get(m, [None, None])[1])

        unknown = [m for m in df['merchant'].unique() if m and m not in cat_mgr.category_map]

        flat_choices = [
            (cat, sub)
            for cat, subs in cat_mgr.valid_categories.items()
            for sub in subs
        ]

        total_unknown = len(unknown)
        for idx, merchant in enumerate(unknown, start=1):
            if self._should_stop:
                return df

            self.log_message.emit('INFO', f'Unknown merchant: {merchant}')

            # Get sample transaction data for this merchant
            merchant_rows = df[df['merchant'] == merchant]
            sample_data = {}
            if not merchant_rows.empty:
                sample_row = merchant_rows.iloc[0]
                if 'amount' in sample_row:
                    sample_data['amount'] = float(sample_row['amount'])
                if 'transaction_date' in sample_row and pd.notna(sample_row['transaction_date']):
                    sample_data['date'] = str(sample_row['transaction_date'])

            # Find similar merchant for suggestion
            # suggested_category = cat_mgr.find_similar_merchant(merchant)

            # Progress text
            progress_text = f"{idx} of {total_unknown} merchants remaining"

            # Pause timeout while waiting for user
            self.stop_timeout()

            # Signal to GUI that we need category selection
            self.category_needed.emit(merchant, flat_choices, sample_data, progress_text)

            # Wait for response
            while not self.response_ready:
                if self._should_stop:
                    return df
                self.msleep(100)

            # Restart timeout
            self.start_timeout()

            if self.category_response:
                cat, sub = self.category_response
                cat_mgr.category_map[merchant] = [cat, sub]
                df.loc[df['merchant'] == merchant, 'category'] = cat
                df.loc[df['merchant'] == merchant, 'subcat'] = sub
                self.log_message.emit('INFO', f'Mapped {merchant} -> {cat}/{sub}')

            self.response_ready = False
            self.category_response = None

        cat_mgr.save_categories()
        return df


class CategoryDialog(QDialog):
    """Dialog for selecting category for unknown merchants with cascading dropdowns."""

    def __init__(self, merchant: str, choices: List[tuple], translations: Translations,
                 parent=None, sample_data: Optional[dict] = None,
                 suggested_category: Optional[tuple] = None,
                 progress_text: Optional[str] = None):
        """
        Initialize category selection dialog with cascading category/subcategory dropdowns.

        Args:
            merchant: Name of the merchant requiring category assignment
            choices: List of (category, subcategory) tuples to choose from
            translations: Translations object for localized UI text
            parent: Parent widget (optional)
            sample_data: Optional dict with sample transaction info (amount, date)
            suggested_category: Optional (category, subcategory) tuple to pre-select
            progress_text: Optional progress text like "3 of 15 merchants remaining"
        """
        super().__init__(parent)
        self.merchant = merchant
        self.choices = choices
        self.translations = translations
        self.selected_category = None

        # Build category structure from choices
        self.category_structure = {}
        for cat, sub in choices:
            if cat not in self.category_structure:
                self.category_structure[cat] = []
            if sub not in self.category_structure[cat]:
                self.category_structure[cat].append(sub)

        self.setWindowTitle(self.translations.get('category_dialog_title'))
        self.setModal(True)
        self.resize(500, 400)

        layout = QVBoxLayout()
        layout.setSpacing(15)

        # Merchant label
        merchant_label = QLabel(self.translations.get('new_merchant', merchant=merchant))
        merchant_label.setFont(QFont('Arial', 12, QFont.Weight.Bold))
        layout.addWidget(merchant_label)

        # Show sample transaction info if available
        if sample_data:
            sample_text = ""
            if 'amount' in sample_data:
                sample_text += f"Amount: ₪{sample_data['amount']:,.2f}"
            if 'date' in sample_data:
                if sample_text:
                    sample_text += "  |  "
                sample_text += f"Date: {sample_data['date']}"
            if sample_text:
                sample_label = QLabel(sample_text)
                sample_label.setFont(QFont('Arial', 9))
                layout.addWidget(sample_label)

        # Category dropdown
        category_label = QLabel(self.translations.get('category') + ":")
        layout.addWidget(category_label)

        self.category_combo = QComboBox()
        self.category_combo.setEditable(True)  # Enable search/filter
        self.category_combo.lineEdit().setPlaceholderText(self.translations.get('search') or "Search...")

        # Add categories sorted alphabetically
        sorted_categories = sorted(self.category_structure.keys())
        for cat in sorted_categories:
            self.category_combo.addItem(cat, cat)

        # Connect category change to update subcategory dropdown
        self.category_combo.currentIndexChanged.connect(self.on_category_changed)
        self.category_combo.lineEdit().textChanged.connect(self.filter_categories)
        layout.addWidget(self.category_combo)

        # Subcategory dropdown
        subcategory_label = QLabel(self.translations.get('subcategory') + ":")
        layout.addWidget(subcategory_label)

        self.subcategory_combo = QComboBox()
        layout.addWidget(self.subcategory_combo)

        # Initialize subcategory dropdown with first category's subcategories
        if sorted_categories:
            self.update_subcategories(sorted_categories[0])

        # Progress text if provided
        if progress_text:
            progress_label = QLabel(progress_text)
            progress_label.setFont(QFont('Arial', 9))
            progress_label.setStyleSheet("color: gray;")
            layout.addWidget(progress_label)

        layout.addStretch()

        # Buttons
        button_layout = QHBoxLayout()

        # Skip button
        skip_btn = QPushButton(self.translations.get('skip') or "Skip")
        skip_btn.clicked.connect(self.skip_merchant)
        button_layout.addWidget(skip_btn)

        button_layout.addStretch()

        # OK/Cancel buttons
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        button_layout.addWidget(buttons)

        layout.addLayout(button_layout)

        # Set default button and enable Enter key
        buttons.button(QDialogButtonBox.StandardButton.Ok).setDefault(True)
        buttons.button(QDialogButtonBox.StandardButton.Ok).setAutoDefault(True)

        self.setLayout(layout)

        # Pre-select suggested category if provided
        if suggested_category:
            cat, sub = suggested_category
            cat_index = self.category_combo.findData(cat)
            if cat_index >= 0:
                self.category_combo.setCurrentIndex(cat_index)
                self.update_subcategories(cat)
                sub_index = self.subcategory_combo.findData(sub)
                if sub_index >= 0:
                    self.subcategory_combo.setCurrentIndex(sub_index)

    def filter_categories(self, text: str):
        """Filter category dropdown based on search text."""
        for i in range(self.category_combo.count()):
            item_text = self.category_combo.itemText(i)
            self.category_combo.setItemData(i, item_text)  # Store original text
            _match = text.lower() in item_text.lower() if text else True
            # Note: QComboBox doesn't support hiding items directly
            # The editable combo will filter as user types

    def on_category_changed(self, index: int):
        """Update subcategory dropdown when category selection changes."""
        if index >= 0:
            category = self.category_combo.itemData(index)
            if category:
                self.update_subcategories(category)

    def update_subcategories(self, category: str):
        """Update subcategory dropdown with subcategories for the given category."""
        self.subcategory_combo.clear()
        if category in self.category_structure:
            subcategories = sorted(self.category_structure[category])
            for sub in subcategories:
                self.subcategory_combo.addItem(sub, sub)

    def skip_merchant(self):
        """Skip this merchant (reject dialog)."""
        self.selected_category = None
        self.reject()

    def accept(self):
        """
        Handle dialog acceptance.

        Stores the selected category/subcategory before closing the dialog.
        """
        cat_index = self.category_combo.currentIndex()
        sub_index = self.subcategory_combo.currentIndex()

        if cat_index >= 0 and sub_index >= 0:
            category = self.category_combo.itemData(cat_index)
            subcategory = self.subcategory_combo.itemData(sub_index)
            if category and subcategory:
                self.selected_category = (category, subcategory)

        if not self.selected_category:
            # Fallback: try to get from combo text if data is missing
            cat_text = self.category_combo.currentText()
            sub_text = self.subcategory_combo.currentText()
            if cat_text and sub_text:
                self.selected_category = (cat_text, sub_text)

        super().accept()


class ConflictDialog(QDialog):
    """Dialog for resolving data conflicts."""

    def __init__(self, month_key: str, translations: Translations, parent=None):
        """
        Initialize conflict resolution dialog.

        Args:
            month_key: Month identifier string (e.g., "2024-01")
            translations: Translations object for localized UI text
            parent: Parent widget (optional)
        """
        super().__init__(parent)
        self.month_key = month_key
        self.translations = translations
        self.decision = "skip"

        self.setWindowTitle(self.translations.get('conflict_title'))
        self.setModal(True)
        self.resize(400, 200)

        layout = QVBoxLayout()

        # Message
        msg = QLabel(self.translations.get('conflict_message', month_key=month_key))
        msg.setFont(QFont('Arial', 11))
        layout.addWidget(msg)

        # Buttons
        override_btn = QPushButton(self.translations.get('override_btn'))
        override_btn.clicked.connect(lambda: self.set_decision("override"))
        layout.addWidget(override_btn)

        add_btn = QPushButton(self.translations.get('add_btn'))
        add_btn.clicked.connect(lambda: self.set_decision("add"))
        layout.addWidget(add_btn)

        skip_btn = QPushButton(self.translations.get('skip_btn'))
        skip_btn.clicked.connect(lambda: self.set_decision("skip"))
        layout.addWidget(skip_btn)

        self.setLayout(layout)

    def set_decision(self, decision: str):
        """
        Set the conflict resolution decision.

        Args:
            decision: One of "override", "add", or "skip"
        """
        self.decision = decision
        self.accept()


class LogViewerWidget(QWidget):
    """Widget for displaying logs with adjustable log level."""

    def __init__(self, translations: Translations, parent=None):
        """
        Initialize log viewer widget.

        Args:
            translations: Translations object
            parent: Parent widget (optional)
        """
        super().__init__(parent)
        self.translations = translations

        from src.config import get_log_level_name

        layout = QVBoxLayout()

        # Controls row: log level selector + clear button
        controls = QHBoxLayout()

        level_label = QLabel(self.translations.get('log_level', 'Log Level') + ':')
        controls.addWidget(level_label)

        self.log_level_combo = QComboBox()
        for level_name in ['DEBUG', 'INFO', 'WARNING', 'ERROR', 'CRITICAL']:
            self.log_level_combo.addItem(level_name)
        self.log_level_combo.setCurrentText(get_log_level_name())
        self.log_level_combo.currentTextChanged.connect(self.on_log_level_changed)
        controls.addWidget(self.log_level_combo)

        controls.addStretch()

        clear_btn = QPushButton(self.translations.get('clear', 'Clear') + ' Logs')
        clear_btn.clicked.connect(self.clear_logs)
        controls.addWidget(clear_btn)

        layout.addLayout(controls)

        # Text area for logs
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setFont(QFont('Consolas', 9))
        layout.addWidget(self.log_text)

        self.setLayout(layout)

    def add_log(self, level: str, message: str):
        """
        Add a log message to the viewer.

        Args:
            level: Log level ('ERROR', 'WARNING', 'INFO', etc.)
            message: Log message text
        """
        timestamp = datetime.now().strftime('%H:%M:%S')

        # Check if message should be displayed based on current log level
        from src.config import get_log_level
        import logging

        level_map = {
            'DEBUG': logging.DEBUG,
            'INFO': logging.INFO,
            'WARNING': logging.WARNING,
            'ERROR': logging.ERROR,
            'CRITICAL': logging.CRITICAL
        }

        current_log_level = get_log_level()
        message_level = level_map.get(level.upper(), logging.INFO)

        # Only show messages at or above current log level
        if message_level < current_log_level:
            return

        # Color code by level
        if level == 'ERROR':
            color = '#ff4444'
        elif level == 'WARNING':
            color = '#ffaa00'
        elif level == 'INFO':
            color = '#4444ff'
        elif level == 'DEBUG':
            color = '#888888'
        else:
            color = '#000000'

        formatted = f"<span style='color: gray;'>[{timestamp}]</span> <span style='color: {color}; font-weight: bold;'>[{level}]</span> {message}<br>"
        self.log_text.insertHtml(formatted)

        # Auto-scroll to bottom
        scrollbar = self.log_text.verticalScrollBar()
        scrollbar.setValue(scrollbar.maximum())

    def clear_logs(self):
        """Clear all log messages from the viewer."""
        self.log_text.clear()

    def on_log_level_changed(self, level_name: str):
        """Update log level setting and logger when user changes selection."""
        import logging
        from src.config import set_log_level, LOG_LEVEL_MAP

        set_log_level(level_name)
        logging.getLogger().setLevel(LOG_LEVEL_MAP[level_name])
        for handler in logging.getLogger().handlers:
            handler.setLevel(LOG_LEVEL_MAP[level_name])


class LogViewerHandler(logging.Handler, QObject):
    """
    Logging handler that forwards log records to the GUI log viewer.
    """
    log_signal = pyqtSignal(str, str)

    def __init__(self, viewer: LogViewerWidget):
        QObject.__init__(self)
        logging.Handler.__init__(self)
        self.log_signal.connect(viewer.add_log)

    def emit(self, record: logging.LogRecord):
        try:
            msg = self.format(record)
            self.log_signal.emit(record.levelname, msg)
        except Exception:
            self.handleError(record)


class QuickStatsWidget(QWidget):
    """Widget displaying quick statistics cards."""

    def __init__(self, translations: Translations, parent=None):
        """
        Initialize quick stats widget.

        Args:
            translations: Translations object for localized UI text
            parent: Parent widget (optional)
        """
        super().__init__(parent)
        self.translations = translations

        layout = QHBoxLayout()
        layout.setSpacing(10)

        # Create 4 stat cards and store value label references
        self.total_card, self.total_value = self.create_stat_card(self.translations.get('total_spending', 'Total Spending'), "₪0.00")
        self.avg_card, self.avg_value = self.create_stat_card(self.translations.get('avg_monthly', 'Avg Monthly'), "₪0.00")
        self.top_category_card, self.top_category_value = self.create_stat_card(self.translations.get('top_category_label', 'Top Category'), "-")
        self.change_card, self.change_value = self.create_stat_card(self.translations.get('vs_average', 'vs Avg'), "0%")

        layout.addWidget(self.total_card)
        layout.addWidget(self.avg_card)
        layout.addWidget(self.top_category_card)
        layout.addWidget(self.change_card)

        layout.addStretch()
        self.setLayout(layout)

    def create_stat_card(self, label: str, value: str) -> tuple:
        """Create a stat card widget and return (card, value_label)."""
        card = QWidget()
        card.setStyleSheet("""
            QWidget {
                background-color: transparent;
                border: 1px solid #DDDDDD;
                border-radius: 8px;
                padding: 10px;
                min-width: 150px;
            }}
        """)

        card_layout = QVBoxLayout()
        card_layout.setSpacing(5)

        label_widget = QLabel(label)
        label_widget.setFont(QFont('Arial', 9))
        card_layout.addWidget(label_widget)

        value_widget = QLabel(value)
        value_widget.setFont(QFont('Arial', 14, QFont.Weight.Bold))
        card_layout.addWidget(value_widget)

        card.setLayout(card_layout)
        return (card, value_widget)

    def update_stats(self, summary_df: pd.DataFrame):
        """
        Update statistics from summary DataFrame.

        Args:
            summary_df: DataFrame with columns: year, month, category, subcat, monthly_amount
        """
        if summary_df.empty:
            self.total_value.setText("₪0.00")
            self.avg_value.setText("₪0.00")
            self.top_category_value.setText("-")
            self.change_value.setText(self.translations.get('not_available', 'N/A'))
            return

        # Total spending
        total = summary_df['monthly_amount'].sum()
        self.total_value.setText(f"₪{total:,.2f}")

        # Average monthly spending
        monthly_totals = summary_df.groupby(['year', 'month'])['monthly_amount'].sum()
        avg_monthly = monthly_totals.mean() if len(monthly_totals) > 0 else 0
        self.avg_value.setText(f"₪{avg_monthly:,.2f}")

        # Top category
        category_totals = summary_df.groupby('category')['monthly_amount'].sum()
        if len(category_totals) > 0:
            top_category = category_totals.idxmax()
            top_amount = category_totals.max()
            self.top_category_value.setText(f"{top_category}\n₪{top_amount:,.2f}")
        else:
            self.top_category_value.setText("-")

        # Change vs average monthly spending
        if len(monthly_totals) >= 1:
            current_month = monthly_totals.iloc[-1]
            avg_monthly_val = monthly_totals.mean()
            if avg_monthly_val > 0:
                change_pct = ((current_month - avg_monthly_val) / avg_monthly_val) * 100
                change_text = f"{change_pct:+.1f}%"
            else:
                change_text = self.translations.get('not_available', 'N/A')
            self.change_value.setText(change_text)
        else:
            self.change_value.setText(self.translations.get('not_available', 'N/A'))


class ChartWidget(QWidget):
    """Widget for displaying matplotlib charts."""

    def __init__(self, translations: Translations, parent=None):
        """
        Initialize chart widget.

        Args:
            translations: Translations object for localized UI text
            parent: Parent widget (optional)
        """
        super().__init__(parent)
        self.translations = translations

        layout = QVBoxLayout()

        # Button layout
        button_layout = QHBoxLayout()

        # Back button (initially hidden) to return to all categories
        self.show_all_btn = QPushButton(self.translations.get('back') or "Back")
        self.show_all_btn.setVisible(False)
        self.show_all_btn.clicked.connect(self.on_show_all_clicked)
        button_layout.addWidget(self.show_all_btn)

        button_layout.addStretch()

        # Export button
        export_btn = QPushButton(self.translations.get('export_chart'))
        export_btn.clicked.connect(self.export_chart)
        button_layout.addWidget(export_btn)

        layout.addLayout(button_layout)

        # Create matplotlib figure
        self.figure = Figure(figsize=(8, 6))
        self.canvas = FigureCanvas(self.figure)
        layout.addWidget(self.canvas)

        self.setLayout(layout)
        self.summary_df = None
        self.full_summary_df = None  # Store full data for "Show All"
        self.on_show_all_callback = None  # Callback to restore full view

    def update_chart(self, summary_df: pd.DataFrame, category_filter: Optional[str] = None,
                     full_summary_df: Optional[pd.DataFrame] = None,
                     show_all_callback: Optional[Callable] = None):
        """
        Update chart with new summary data.

        Creates a bar chart showing monthly spending totals. Handles RTL languages
        (Hebrew) by reversing axis and text rendering. Can filter by category.

        Args:
            summary_df: DataFrame with columns: year, month, category, subcat, monthly_amount
            category_filter: Optional category name to filter by
            full_summary_df: Optional full summary DataFrame for "Show All" functionality
            show_all_callback: Optional callback function to call when "Show All" is clicked
        """
        try:
            self.summary_df = summary_df
            self.full_summary_df = full_summary_df if full_summary_df is not None else summary_df
            self.on_show_all_callback = show_all_callback
            self.figure.clear()

            # Show/hide "Show All" button based on whether we're filtering
            self.show_all_btn.setVisible(category_filter is not None and category_filter != "")

            # Filter by category if specified
            if category_filter and not summary_df.empty:
                filtered_df = summary_df[summary_df['category'] == category_filter]
                if not filtered_df.empty:
                    # Group by month and subcategory for filtered view
                    monthly_data = filtered_df.groupby('month')['monthly_amount'].sum()
                else:
                    monthly_data = pd.Series(dtype=float)
            elif not summary_df.empty:
                # Group by month and sum amounts (if we have data)
                monthly_data = summary_df.groupby('month')['monthly_amount'].sum()
            else:
                monthly_data = pd.Series(dtype=float)

            # Create data for all 12 months (1-12)
            all_months = range(1, 13)
            values = [monthly_data.get(m, 0.0) for m in all_months]

            ax = self.figure.add_subplot(111)

            # Get month names in current language for all 12 months
            months = [self.translations.get(f'month_{m}') for m in all_months]

            # Get labels in current language
            if category_filter:
                title = self.translations.get('year_overview_category', default=f"{self.translations.get('year_overview')} - {category_filter}", category=category_filter)
            else:
                title = self.translations.get('year_overview')
            xlabel = self.translations.get('month')
            ylabel = self.translations.get('amount')

            # For RTL languages (Hebrew), we need to handle both direction and text rendering
            if self.translations.is_rtl():
                # 1. Reverse the order of data for RTL display (Dec -> Jan)
                months = months[::-1]
                values = values[::-1]

                # 2. Fix text rendering by reversing strings (Matplotlib renders Hebrew LTR by default)
                # This makes "שלום" appear correctly as "שלום" instead of "םולש" (or vice versa depending on font)
                title = title[::-1]
                xlabel = xlabel[::-1]
                ylabel = ylabel[::-1]
                months = [m[::-1] for m in months]

            bars = ax.bar(months, values, color='#4CAF50')

            # Add value labels on top of bars
            for bar, val in zip(bars, values):
                label = f"₪{val:,.0f}" if val != 0 else ""
                if label:
                    ax.annotate(
                        label,
                        xy=(bar.get_x() + bar.get_width() / 2, bar.get_height()),
                        xytext=(0, 4),
                        textcoords="offset points",
                        ha='center',
                        va='bottom',
                        fontsize=8,
                        color='#333333'
                    )

            ax.set_xlabel(xlabel)
            ax.set_ylabel(ylabel)
            ax.set_title(title)

            # Rotate labels for better readability
            ax.tick_params(axis='x', rotation=45)

            # For RTL, invert x-axis so it goes right to left
            if self.translations.is_rtl():
                ax.invert_xaxis()

            self.canvas.draw()
        except Exception as e:
            logger = logging.getLogger(__name__)
            logger.error(f"Chart update error: {e}", exc_info=True)

    def on_show_all_clicked(self):
        """Handle click on 'Show All Categories' button."""
        if self.on_show_all_callback:
            self.on_show_all_callback()
        elif self.full_summary_df is not None:
            self.update_chart(self.full_summary_df, category_filter=None)

    def export_chart(self):
        """
        Export the current chart as an image file.

        Opens a file dialog to save the chart as PNG or PDF. Shows success message
        when export completes.
        """
        if self.summary_df is None:
            return

        filename, _ = QFileDialog.getSaveFileName(
            self,
            self.translations.get('export_chart'),
            f"budget_chart_{datetime.now().strftime('%Y%m%d')}.png",
            "PNG Files (*.png);;PDF Files (*.pdf)"
        )

        if filename:
            self.figure.savefig(filename, dpi=300, bbox_inches='tight')
            QMessageBox.information(
                self,
                self.translations.get('success_title'),
                self.translations.get('chart_exported')
            )


class FileListWidget(QListWidget):
    """Custom list widget with drag-and-drop support."""

    files_dropped = pyqtSignal(list)

    def __init__(self, parent=None):
        """
        Initialize file list widget with drag-and-drop support.

        Args:
            parent: Parent widget (optional)
        """
        super().__init__(parent)
        self.setAcceptDrops(True)

    def dragEnterEvent(self, event: QDragEnterEvent):
        """
        Handle drag enter event.

        Args:
            event: Drag enter event
        """
        if event.mimeData().hasUrls():
            event.acceptProposedAction()

    def dropEvent(self, event: QDropEvent):
        """
        Handle file drop event.

        Extracts Excel file paths from dropped URLs and emits files_dropped signal.

        Args:
            event: Drop event
        """
        files = []
        for url in event.mimeData().urls():
            file_path = url.toLocalFile()
            if file_path.endswith(('.xlsx', '.xls')):
                files.append(file_path)

        if files:
            self.files_dropped.emit(files)



class CategoryManagementDialog(QDialog):
    """Dialog for managing category mappings."""

    def __init__(self, category_manager: CategoryManager, translations: Translations, parent=None):
        """
        Initialize category management dialog.

        Args:
            category_manager: CategoryManager instance with current mappings
            translations: Translations object for localized UI text
            parent: Parent widget (optional)
        """
        super().__init__(parent)
        self.category_manager = category_manager
        self.translations = translations
        self.original_mappings = category_manager.category_map.copy()

        title = self.translations.get('category_management_title') or 'Category Management'
        self.setWindowTitle(title)
        self.setModal(True)
        self.resize(800, 600)

        layout = QVBoxLayout()

        # Search/filter box
        search_layout = QHBoxLayout()
        search_label = QLabel(self.translations.get('search') or 'Search:')
        search_layout.addWidget(search_label)

        self.search_box = QLineEdit()
        self.search_box.setPlaceholderText(self.translations.get('search_merchants') or 'Search merchants...')
        self.search_box.textChanged.connect(self.filter_table)
        # Enable Ctrl+F to focus search
        from PyQt5.QtWidgets import QShortcut
        from PyQt5.QtGui import QKeySequence
        search_shortcut = QShortcut(QKeySequence('Ctrl+F'), self)
        search_shortcut.activated.connect(lambda: self.search_box.setFocus())
        search_layout.addWidget(self.search_box)

        clear_search_btn = QPushButton(self.translations.get('clear') or 'Clear')
        clear_search_btn.clicked.connect(self.clear_search)
        search_layout.addWidget(clear_search_btn)

        layout.addLayout(search_layout)

        # Table for merchant mappings
        self.table = QTableWidget()
        self.table.setColumnCount(3)
        self.table.setHorizontalHeaderLabels([
            self.translations.get('merchant') or 'Merchant',
            self.translations.get('category') or 'Category',
            self.translations.get('subcategory') or 'Subcategory'
        ])
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        layout.addWidget(self.table)

        # Info label
        self.info_label = QLabel()
        layout.addWidget(self.info_label)

        # Buttons
        button_layout = QHBoxLayout()

        edit_btn = QPushButton(self.translations.get('edit') or 'Edit')
        edit_btn.clicked.connect(self.edit_mapping)
        button_layout.addWidget(edit_btn)

        delete_btn = QPushButton(self.translations.get('delete') or 'Delete')
        delete_btn.clicked.connect(self.delete_mapping)
        button_layout.addWidget(delete_btn)

        button_layout.addStretch()

        close_btn = QPushButton(self.translations.get('close') or 'Close')
        close_btn.clicked.connect(self.accept)
        button_layout.addWidget(close_btn)

        layout.addLayout(button_layout)
        self.setLayout(layout)

        # Populate table
        self.populate_table()

    def populate_table(self):
        """Populate table with current category mappings."""
        self.table.setRowCount(0)

        for merchant, (category, subcat) in self.category_manager.category_map.items():
            row = self.table.rowCount()
            self.table.insertRow(row)

            self.table.setItem(row, 0, QTableWidgetItem(str(merchant)))
            self.table.setItem(row, 1, QTableWidgetItem(str(category)))
            self.table.setItem(row, 2, QTableWidgetItem(str(subcat)))

        self.update_info_label()
        self.table.resizeColumnsToContents()

    def filter_table(self):
        """Filter table based on search box text."""
        search_text = self.search_box.text().lower()

        for row in range(self.table.rowCount()):
            merchant = self.table.item(row, 0).text().lower()
            category = self.table.item(row, 1).text().lower()
            subcat = self.table.item(row, 2).text().lower()

            match = (search_text in merchant or
                    search_text in category or
                    search_text in subcat)

            self.table.setRowHidden(row, not match)

        self.update_info_label()

    def clear_search(self):
        """Clear search box and show all rows."""
        self.search_box.clear()
        for row in range(self.table.rowCount()):
            self.table.setRowHidden(row, False)
        self.update_info_label()

    def update_info_label(self):
        """Update info label with count of visible/total merchants."""
        total = self.table.rowCount()
        visible = sum(1 for row in range(total) if not self.table.isRowHidden(row))

        if visible == total:
            self.info_label.setText(f"Total: {total} merchant(s)")
        else:
            self.info_label.setText(f"Showing: {visible} of {total} merchant(s)")

    def get_selected_merchant(self) -> Optional[str]:
        """Get selected merchant from table."""
        current_row = self.table.currentRow()
        if current_row >= 0:
            return self.table.item(current_row, 0).text()
        return None

    def edit_mapping(self):
        """Edit selected merchant's category mapping."""
        merchant = self.get_selected_merchant()
        if not merchant:
            QMessageBox.warning(self, self.translations.get('warning_title'),
                              "Please select a merchant to edit.")
            return

        # current_cat = self.category_manager.category_map.get(merchant, [None, None])

        # Get all valid category choices
        flat_choices = [
            (cat, sub)
            for cat, subs in self.category_manager.valid_categories.items()
            for sub in subs
        ]

        dialog = CategoryDialog(merchant, flat_choices, self.translations, self)
        if dialog.exec() == QDialog.DialogCode.Accepted and dialog.selected_category:
            cat, sub = dialog.selected_category
            self.category_manager.category_map[merchant] = [cat, sub]
            self.category_manager.save_categories()
            self.populate_table()
            QMessageBox.information(self, self.translations.get('success_title'),
                                  f"Updated mapping for {merchant}")

    def delete_mapping(self):
        """Delete selected merchant's category mapping."""
        merchant = self.get_selected_merchant()
        if not merchant:
            QMessageBox.warning(self, self.translations.get('warning_title'),
                              "Please select a merchant to delete.")
            return

        reply = QMessageBox.question(
            self,
            self.translations.get('confirm_title') or 'Confirm',
            f"Delete mapping for '{merchant}'?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )

        if reply == QMessageBox.StandardButton.Yes:
            del self.category_manager.category_map[merchant]
            self.category_manager.save_categories()
            self.populate_table()
            QMessageBox.information(self, self.translations.get('success_title'),
                                  f"Deleted mapping for {merchant}")


class BudgetTrackerGUI(QMainWindow):
    """Main application window."""

    def __init__(self, language='he'):
        """
        Initialize the main application window.

        Args:
            language: Initial language code ('he' for Hebrew, 'en' for English)
        """
        super().__init__()

        # Set application icon (works for both dev and PyInstaller)
        if getattr(sys, 'frozen', False):
            # Running as bundled exe
            base_path = Path(sys._MEIPASS)
        else:
            # Running in development
            base_path = Path(__file__).parent
        icon_path = base_path / "badget_tracker.ico"
        if icon_path.exists():
            self.setWindowIcon(QIcon(str(icon_path)))

        # Initialize translations
        self.translations = Translations(language)

        # Setup logging
        ensure_dirs([OUTPUT_DIR])
        from src.config import get_log_level
        setup_logging(OUTPUT_DIR, LOG_FILE_NAME, log_level=get_log_level())

        # Initialize UI
        self.init_ui()

        # Add keyboard shortcuts
        self.setup_shortcuts()

        # Load initial file list
        self.refresh_files()

    def focus_search_if_available(self):
        """Focus search box if category management dialog is open."""
        # This will be handled by the dialog itself if needed
        pass

    def _attach_log_handler(self):
        """Attach a logging handler that mirrors logs into the Logs tab."""
        try:
            handler = LogViewerHandler(self.log_viewer)
            handler.setLevel(get_log_level())
            handler.setFormatter(logging.Formatter('%(asctime)s [%(levelname)s] %(message)s'))
            logging.getLogger().addHandler(handler)
            logging.getLogger().setLevel(get_log_level())
            self._log_viewer_handler = handler
        except Exception as e:
            # Fallback to console if we cannot attach handler
            print(f"Failed to attach log handler: {e}")

    def handle_escape(self):
        """Handle Escape key - close dialogs or clear selections."""
        # Close any open dialogs
        for widget in QApplication.topLevelWidgets():
            if isinstance(widget, QDialog) and widget.isVisible():
                widget.reject()
                return

        # Clear table selection
        if hasattr(self, 'preview_table'):
            self.preview_table.clearSelection()
            if hasattr(self, 'current_category_filter'):
                self.current_category_filter = None
                if hasattr(self, 'last_summary_df'):
                    self.chart_widget.update_chart(self.last_summary_df)
                    self.statusBar().showMessage("Selection cleared")

    def setup_shortcuts(self):
        """Setup keyboard shortcuts."""
        from PyQt5.QtWidgets import QShortcut
        from PyQt5.QtGui import QKeySequence

        # F5 - Refresh files
        refresh_shortcut = QShortcut(QKeySequence('F5'), self)
        refresh_shortcut.activated.connect(self.refresh_files)

        # Ctrl+O - Open dashboard
        open_shortcut = QShortcut(QKeySequence('Ctrl+O'), self)
        open_shortcut.activated.connect(self.open_dashboard)

        # Ctrl+R - Reload dashboard data
        reload_shortcut = QShortcut(QKeySequence('Ctrl+R'), self)
        reload_shortcut.activated.connect(self.load_dashboard_data)

        # Ctrl+P - Process transactions
        process_shortcut = QShortcut(QKeySequence('Ctrl+P'), self)
        process_shortcut.activated.connect(self.process_transactions)

        # Ctrl+F - Focus search (when category management is open)
        search_shortcut = QShortcut(QKeySequence('Ctrl+F'), self)
        search_shortcut.activated.connect(self.focus_search_if_available)

        # Esc - Close dialogs or clear selections
        esc_shortcut = QShortcut(QKeySequence('Escape'), self)
        esc_shortcut.activated.connect(self.handle_escape)

    def init_ui(self):
        """
        Initialize and set up the user interface.

        Creates all UI components including file management section, preview tabs
        (chart, table, logs, archive), action buttons, and status bar. Sets RTL
        layout for Hebrew language. Loads dashboard data and performs startup validation.
        """
        self.setWindowTitle(self.translations.get('app_title'))
        self.resize(1200, 800)

        # Set RTL layout for Hebrew
        if self.translations.is_rtl():
            self.setLayoutDirection(Qt.LayoutDirection.RightToLeft)

        # Central widget
        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        main_layout = QVBoxLayout()
        central_widget.setLayout(main_layout)

        # Header with language toggle
        header_layout = QHBoxLayout()

        header_layout.addStretch() # Push title to center

        title_label = QLabel(self.translations.get('app_title'))
        title_label.setFont(QFont('Arial', 16, QFont.Weight.Bold))
        header_layout.addWidget(title_label)

        header_layout.addStretch()

        self.lang_toggle = QPushButton('EN / עב')
        self.lang_toggle.clicked.connect(self.toggle_language)
        header_layout.addWidget(self.lang_toggle)

        main_layout.addLayout(header_layout)

        # File management section
        file_group = QWidget()
        file_layout = QVBoxLayout()
        file_group.setLayout(file_layout)

        file_label = QLabel(self.translations.get('file_management'))
        file_label.setFont(QFont('Arial', 12, QFont.Weight.Bold))
        file_layout.addWidget(file_label)

        self.file_list = FileListWidget()
        self.file_list.files_dropped.connect(self.import_dropped_files)
        file_layout.addWidget(self.file_list)

        self.file_info_label = QLabel()
        file_layout.addWidget(self.file_info_label)

        file_buttons = QHBoxLayout()
        import_btn = QPushButton(self.translations.get('import_files'))
        import_btn.clicked.connect(self.import_files)
        file_buttons.addWidget(import_btn)

        refresh_all_btn = QPushButton(self.translations.get('refresh_all'))
        refresh_all_btn.setFont(QFont('Arial', 10, QFont.Weight.Bold))
        refresh_all_btn.clicked.connect(self.refresh_all)
        file_buttons.addWidget(refresh_all_btn)

        file_layout.addLayout(file_buttons)
        main_layout.addWidget(file_group)

        self.process_btn = QPushButton(self.translations.get('process_transactions'))
        self.process_btn.setFont(QFont('Arial', 14, QFont.Weight.Bold))
        self.process_btn.setMinimumHeight(50)
        self.process_btn.clicked.connect(self.process_transactions)
        main_layout.addWidget(self.process_btn)

        # Progress bar
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        main_layout.addWidget(self.progress_bar)

        # Quick stats widget
        self.quick_stats = QuickStatsWidget(self.translations)
        main_layout.addWidget(self.quick_stats)

        # Preview tabs
        self.preview_tabs = QTabWidget()

        # Chart tab
        self.chart_widget = ChartWidget(self.translations)
        self.preview_tabs.addTab(self.chart_widget, self.translations.get('chart_tab'))

        # Table tab with back button
        self.table_tab_widget = QWidget()
        table_layout = QVBoxLayout()
        table_layout.setContentsMargins(0, 0, 0, 0)

        table_back_btn = QPushButton(self.translations.get('back'))
        table_back_btn.setFixedWidth(120)
        table_back_btn.clicked.connect(lambda: self.show_all_categories(switch_tab=False, clear_selection=True))
        table_layout.addWidget(table_back_btn, alignment=Qt.AlignmentFlag.AlignLeft)

        self.preview_table = QTableWidget()
        # Table is read-only; selection drives the chart
        self.preview_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.preview_table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        table_layout.addWidget(self.preview_table)

        self.table_tab_widget.setLayout(table_layout)
        self.preview_tabs.addTab(self.table_tab_widget, self.translations.get('table_tab'))

        # Logs tab
        self.log_viewer = LogViewerWidget(self.translations)
        self.preview_tabs.addTab(self.log_viewer, "Logs")
        self._attach_log_handler()

        # Archive tab
        archive_widget = QWidget()
        archive_layout = QVBoxLayout()
        archive_widget.setLayout(archive_layout)

        self.archive_list = QListWidget()
        archive_layout.addWidget(self.archive_list)

        self.archive_info_label = QLabel()
        archive_layout.addWidget(self.archive_info_label)

        clear_archive_btn = QPushButton(self.translations.get('clear_archive'))
        clear_archive_btn.clicked.connect(self.clear_archive)
        archive_layout.addWidget(clear_archive_btn)

        self.preview_tabs.addTab(archive_widget, self.translations.get('archive'))

        main_layout.addWidget(self.preview_tabs)

        # Set Table tab as default
        self.preview_tabs.setCurrentIndex(0)

        # Action buttons
        action_layout = QHBoxLayout()

        open_dashboard_btn = QPushButton(self.translations.get('open_dashboard'))
        open_dashboard_btn.clicked.connect(self.open_dashboard)
        action_layout.addWidget(open_dashboard_btn)

        manage_categories_btn = QPushButton(self.translations.get('manage_categories'))
        manage_categories_btn.clicked.connect(self.show_category_management)
        action_layout.addWidget(manage_categories_btn)

        action_layout.addStretch()

        exit_btn = QPushButton(self.translations.get('exit'))
        exit_btn.clicked.connect(self.close)
        action_layout.addWidget(exit_btn)

        main_layout.addLayout(action_layout)

        # Status bar
        self.statusBar().showMessage(self.translations.get('app_title'))

        # Refresh archive list
        self.refresh_archive()

        # Load existing dashboard data
        self.load_dashboard_data()

        # Perform startup validation (non-blocking)
        QTimer.singleShot(500, self.perform_startup_validation)

    def perform_startup_validation(self):
        """Perform template validation on startup and show warnings if needed."""
        try:
            from src.category_manager import CategoryManager

            # Run validation
            manager = CategoryManager(CATEGORIES_FILE_PATH, DASHBOARD_FILE_PATH, strict_validation=False)
            validation = manager.validate_template_structure()

            # Only show message if there are errors or warnings
            if validation.errors or validation.warnings:
                msg_parts = []

                if validation.errors:
                    msg_parts.append(f"Found {len(validation.errors)} error(s) in template:")
                    for i, error in enumerate(validation.errors[:3], 1):  # Show first 3
                        msg_parts.append(f"  {i}. {error}")
                    if len(validation.errors) > 3:
                        msg_parts.append(f"  ... and {len(validation.errors) - 3} more")

                if validation.warnings:
                    msg_parts.append(f"Found {len(validation.warnings)} warning(s):")
                    for i, warning in enumerate(validation.warnings[:3], 1):
                        msg_parts.append(f"  {i}. {warning}")
                    if len(validation.warnings) > 3:
                        msg_parts.append(f"  ... and {len(validation.warnings) - 3} more")

                # Show appropriate dialog
                if validation.errors:
                    QMessageBox.warning(
                        self,
                        "Template Validation Issues",
                        "\n".join(msg_parts)
                    )

        except Exception as e:
            # Don't block startup for validation errors, just log them
            self.log_viewer.add_log('ERROR', f'Failed to validate template on startup: {str(e)}')

    def import_dropped_files(self, files: List[str]):
        """Import files dropped via drag and drop."""
        from src.validators import validate_excel_file, ValidationError
        count, errors = 0, []

        for file_path in files:
            try:
                validate_excel_file(Path(file_path))
                dest_path = TRANSACTIONS_DIR / Path(file_path).name
                shutil.copy2(file_path, dest_path)
                count += 1
                self.log_viewer.add_log('INFO', f'Imported: {Path(file_path).name}')
            except ValidationError as e:
                errors.append(f"{Path(file_path).name}: {e}")
                self.log_viewer.add_log('ERROR', f'Skipped - {errors[-1]}')

        self.refresh_files()
        if count > 0:
            msg = self.translations.get('files_imported', count=count)
            QMessageBox.information(self, self.translations.get('success_title'), msg)
        if errors:
            QMessageBox.warning(self, "Import Errors", "\\n".join(errors[:5]))

    def load_dashboard_data(self):
        """
        Load existing data from dashboard Excel file.

        Reads all year sheets from the dashboard and displays the data in the preview
        table and chart. Filters to show only the current year's data.
        Validates file integrity and checks for file locks before loading.
        """
        logger = logging.getLogger(__name__)
        logger.debug(f"Loading dashboard data from {DASHBOARD_FILE_PATH}")
        try:
            if not DASHBOARD_FILE_PATH.exists():
                logger.debug("Dashboard file does not exist")
                self.log_viewer.add_log('INFO', 'No existing dashboard found')
                return

            # Validate dashboard integrity
            is_valid, error_msg = validate_dashboard_integrity(DASHBOARD_FILE_PATH)
            if not is_valid:
                self.log_viewer.add_log('WARNING', f'Dashboard validation failed: {error_msg}')
                QMessageBox.warning(
                    self,
                    self.translations.get('warning_title'),
                    f"Dashboard file issue detected:\n\n{error_msg}\n\nPlease check the file and try again."
                )
                return

            self.log_viewer.add_log('INFO', 'Loading existing dashboard data...')
            logger.debug("File exists, loading...")

            # Read all sheets except template
            import openpyxl
            wb = openpyxl.load_workbook(DASHBOARD_FILE_PATH, data_only=True)
            logger.debug(f"Workbook loaded. Sheets: {wb.sheetnames}")
            try:
                all_data = []
                for sheet_name in wb.sheetnames:
                    if sheet_name.lower() == 'template':
                        continue

                    ws = wb[sheet_name]

                    # Try to parse as year
                    try:
                        year = int(sheet_name)
                    except ValueError:
                        continue

                    # Read data from sheet
                    row_count = 0

                    # Track last category/subcategory for merged cells
                    last_category = None
                    last_subcat = None

                    # Iterate through rows starting from row 2
                    for row_idx in range(2, ws.max_row + 1):
                        # Read category and subcategory
                        category_val = ws.cell(row=row_idx, column=1).value
                        subcat_val = ws.cell(row=row_idx, column=2).value

                        # Use last value if current is None (for merged cells)
                        if category_val:
                            last_category = category_val
                        if subcat_val:
                            last_subcat = subcat_val

                        category = last_category
                        subcat = last_subcat

                        # Skip summary rows to avoid double-counting
                        if not category or not subcat or 'סיכום' in str(category):
                            continue

                        # Read each month (columns 3-14 = C-N = Jan-Dec)
                        for month in range(1, 13):
                            col_idx = month + 2  # month 1 = column 3 (C), month 12 = column 14 (N)
                            amount = ws.cell(row=row_idx, column=col_idx).value

                            if amount and isinstance(amount, (int, float)) and amount != 0:
                                all_data.append({
                                    'year': year,
                                    'month': month,
                                    'category': str(category),
                                    'subcat': str(subcat),
                                    'monthly_amount': float(amount)
                                })
                                row_count += 1
            finally:
                wb.close()

            total_amount = sum(d['monthly_amount'] for d in all_data)
            logger.debug(f"Total records found: {len(all_data)}")
            logger.debug(f"Total amount loaded: {total_amount:,.2f}")
            if all_data:
                summary_df = pd.DataFrame(all_data)

                # Filter to current year only
                import datetime
                current_year = datetime.datetime.now().year
                summary_df = summary_df[summary_df['year'] == current_year]

                if not summary_df.empty:
                    self.log_viewer.add_log('INFO', f'Loaded {len(summary_df)} records from dashboard ({current_year})')

                    # Update displays
                    self.update_preview_table(summary_df)
                    self.chart_widget.update_chart(summary_df)
                    self.quick_stats.update_stats(summary_df)

                    # Update chart tab title
                    years = sorted(summary_df['year'].unique())
                    year_str = ", ".join(map(str, years))
                    self.preview_tabs.setTabText(self.preview_tabs.indexOf(self.chart_widget),
                                                 f"{self.translations.get('chart_tab')} ({year_str})")
                    # Keep table tab title in sync with years
                    self.preview_tabs.setTabText(self.preview_tabs.indexOf(self.table_tab_widget),
                                                 f"{self.translations.get('table_tab')} ({year_str})")

                    self.log_viewer.add_log('INFO', 'Dashboard data loaded successfully')
                else:
                    self.log_viewer.add_log('INFO', f'No data found for current year ({current_year})')
            else:
                self.log_viewer.add_log('INFO', 'No data found in dashboard')

        except FileNotFoundError as e:
            error_msg = get_user_friendly_error(e)
            self.log_viewer.add_log('ERROR', f'Failed to load dashboard: {str(e)}')
            QMessageBox.warning(self, self.translations.get('warning_title'), error_msg)
        except PermissionError as e:
            error_msg = get_user_friendly_error(e)
            self.log_viewer.add_log('ERROR', f'Permission error loading dashboard: {str(e)}')
            QMessageBox.warning(self, self.translations.get('warning_title'), error_msg)
        except (IOError, OSError) as e:
            error_msg = get_user_friendly_error(e)
            self.log_viewer.add_log('ERROR', f'IO error loading dashboard: {str(e)}')
            QMessageBox.warning(self, self.translations.get('warning_title'), error_msg)
        except Exception as e:
            error_msg = get_user_friendly_error(e)
            self.log_viewer.add_log('WARNING', f'Failed to load dashboard data: {str(e)}')
            QMessageBox.warning(self, self.translations.get('warning_title'), error_msg)

    def toggle_language(self):
        """
        Toggle between Hebrew and English interface.

        Recreates the entire UI with the new language setting.
        """
        new_lang = 'en' if self.translations.language == 'he' else 'he'
        self.translations.set_language(new_lang)

        # Recreate UI with new language
        self.close()
        self.__init__(new_lang)
        self.show()

    def refresh_files(self):
        """
        Refresh the transaction files list.

        Scans the transactions directory for Excel files and updates the file list
        widget. Displays file count and total size information.
        """
        self.file_list.clear()

        if not TRANSACTIONS_DIR.exists():
            ensure_dirs([TRANSACTIONS_DIR])
            return

        files = list(TRANSACTIONS_DIR.glob('*.xls*'))
        for file_path in files:
            self.file_list.addItem(file_path.name)

        total_size = sum(f.stat().st_size for f in files)
        size_mb = total_size / (1024 * 1024)

        self.file_info_label.setText(
            f"{self.translations.get('files_count', count=len(files))} | "
            f"{self.translations.get('total_size', size=f'{size_mb:.2f} MB')}"
        )

    def refresh_archive(self):
        """
        Refresh the archive files list.

        Scans the archive directory for Excel files and updates the archive list widget.
        Displays archive file count.
        """
        self.archive_list.clear()

        archive_path = ARCHIVE_DIR
        if not archive_path.exists():
            return

        files = list(archive_path.glob('*.xls*'))
        for file_path in files:
            self.archive_list.addItem(file_path.name)

        self.archive_info_label.setText(
            self.translations.get('archive_count', count=len(files))
        )

    def refresh_all(self):
        """
        Refresh everything: files, archive, and dashboard data.

        Recreates the entire UI to refresh all elements including file lists,
        archive, and dashboard data display.
        """
        # Recreate UI to refresh all elements (like language change does)
        self.close()

        self.__init__(self.translations.language)
        self.show()

    def import_files(self):
        """
        Import Excel files via file dialog.

        Opens a file selection dialog allowing user to choose one or more Excel files.
        Copies selected files to the transactions directory and refreshes the file list.
        Shows success message with count of imported files.
        Validates files before importing and provides user-friendly error messages.
        """
        files, _ = QFileDialog.getOpenFileNames(
            self,
            self.translations.get('select_files'),
            "",
            "Excel Files (*.xlsx *.xls)"
        )

        if not files:
            return

        count = 0
        errors = []
        for file_path in files:
            try:
                src = Path(file_path)

                # Check if source file exists
                if not src.exists():
                    errors.append(f"{src.name}: File not found")
                    continue

                # Check file permissions
                has_permission, perm_error = check_file_permissions(src, 'read')
                if not has_permission:
                    errors.append(f"{src.name}: {perm_error}")
                    continue

                # Check if source file is locked
                if is_file_locked(src):
                    errors.append(f"{src.name}: File is locked (may be open in Excel)")
                    continue

                dst = TRANSACTIONS_DIR / src.name

                # Check if destination file is locked
                if dst.exists() and is_file_locked(dst):
                    errors.append(f"{src.name}: Destination file is locked")
                    continue

                shutil.copy2(src, dst)
                count += 1
                self.log_viewer.add_log('INFO', f'Imported file: {src.name}')

            except FileNotFoundError as e:
                error_msg = get_user_friendly_error(e)
                errors.append(f"{Path(file_path).name}: {error_msg}")
            except PermissionError as e:
                error_msg = get_user_friendly_error(e)
                errors.append(f"{Path(file_path).name}: {error_msg}")
            except (IOError, OSError) as e:
                error_msg = get_user_friendly_error(e)
                errors.append(f"{Path(file_path).name}: {error_msg}")
            except Exception as e:
                error_msg = get_user_friendly_error(e)
                errors.append(f"{Path(file_path).name}: {error_msg}")

        self.refresh_files()

        if count > 0:
            QMessageBox.information(
                self,
                self.translations.get('success_title'),
                self.translations.get('files_imported', count=count)
            )

        if errors:
            error_text = "\n".join(errors[:5])  # Show first 5 errors
            if len(errors) > 5:
                error_text += f"\n... and {len(errors) - 5} more errors"
            self.log_viewer.add_log('ERROR', f'Import errors: {len(errors)} file(s) failed')
            QMessageBox.warning(
                self,
                "Import Errors",
                f"Some files could not be imported:\n\n{error_text}"
            )

    def process_transactions(self):
        """
        Process transaction files from the transactions directory.

        Starts a background thread to load, normalize, categorize, and preview
        transactions. Disables the process button and shows progress bar during processing.
        Switches to Logs tab to show progress messages.
        """
        # Disable button during processing
        self.process_btn.setEnabled(False)
        self.progress_bar.setVisible(True)
        self.progress_bar.setRange(0, 0)  # Indeterminate progress

        # Switch to Logs tab to show progress
        self.preview_tabs.setCurrentWidget(self.log_viewer)

        # Create and start processing thread
        self.thread = ProcessThread(self.translations)
        self.thread.progress.connect(self.update_progress)
        self.thread.finished.connect(self.processing_finished)
        self.thread.error.connect(self.processing_error)
        self.thread.category_needed.connect(self.show_category_dialog)
        self.thread.log_message.connect(self.log_viewer.add_log)
        self.thread.start()


    def update_progress(self, message: str):
        """
        Update progress status message.

        Args:
            message: Status message to display in the status bar
        """
        self.statusBar().showMessage(message)

    def resolve_conflict(self, month_key: str) -> str:
        """
        Resolve data conflict via GUI dialog.

        Shows a dialog asking user how to handle conflicting data for a month.

        Args:
            month_key: Month identifier string (e.g., "2024-01")

        Returns:
            User's decision: "override", "add", or "skip"
        """
        dialog = ConflictDialog(month_key, self.translations, self)
        dialog.exec()
        return dialog.decision

    def show_category_dialog(self, merchant: str, choices: List[tuple], sample_data: dict, progress_text: str):
        """
        Show dialog for category selection.

        Displays a modal dialog asking user to assign a category/subcategory to
        an unknown merchant. Updates the processing thread with the user's selection.

        Args:
            merchant: Name of the merchant requiring category assignment
            choices: List of (category, subcategory) tuples to choose from
            sample_data: Dict with sample transaction info (amount, date)
            progress_text: Progress text like "3 of 15 merchants remaining"
        """
        # Get category manager to find similar merchant
        from src.category_manager import CategoryManager
        cat_mgr = CategoryManager(CATEGORIES_FILE_PATH, DASHBOARD_FILE_PATH)
        suggested_category = cat_mgr.find_similar_merchant(merchant)

        dialog = CategoryDialog(
            merchant,
            choices,
            self.translations,
            self,
            sample_data=sample_data,
            suggested_category=suggested_category,
            progress_text=progress_text
        )
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.thread.category_response = dialog.selected_category
        else:
            self.thread.category_response = None

        self.thread.response_ready = True

    def processing_finished(self, summary_df: pd.DataFrame, success: bool):
        """
        Handle processing completion.

        Updates UI with results, writes data to dashboard, refreshes file lists,
        and shows success message.

        Args:
            summary_df: DataFrame with processed transaction summary
            success: Whether processing completed successfully
        """
        self.process_btn.setEnabled(True)
        self.progress_bar.setVisible(False)

        if success:
            # Update preview table
            self.update_preview_table(summary_df)

            # Update chart
            try:
                self.chart_widget.update_chart(summary_df)

                # Update chart tab title
                years = sorted(summary_df['year'].unique())
                year_str = ", ".join(map(str, years))
                self.preview_tabs.setTabText(self.preview_tabs.indexOf(self.chart_widget),
                                             f"{self.translations.get('chart_tab')} ({year_str})")

                self.log_viewer.add_log('INFO', 'Chart updated successfully')
            except Exception as e:
                self.log_viewer.add_log('WARNING', f'Failed to update chart: {str(e)}')

            # Update quick stats
            try:
                self.quick_stats.update_stats(summary_df)
            except Exception as e:
                self.log_viewer.add_log('WARNING', f'Failed to update quick stats: {str(e)}')
            except Exception as e:
                self.log_viewer.add_log('WARNING', f'Failed to update chart: {str(e)}')

            # Write to dashboard
            self.log_viewer.add_log('INFO', 'Writing to dashboard...')
            try:
                # Check if dashboard is locked before writing
                if is_file_locked(DASHBOARD_FILE_PATH):
                    error_msg = get_user_friendly_error(PermissionError("Dashboard file is locked"))
                    QMessageBox.warning(
                        self,
                        self.translations.get('warning_title'),
                        f"Cannot write to dashboard:\n\n{error_msg}"
                    )
                    self.log_viewer.add_log('WARNING', 'Skipped dashboard write - file is locked')
                else:
                    writer = DashboardWriter(DASHBOARD_FILE_PATH)
                    writer.update(summary_df, conflict_resolver=self.resolve_conflict)
                    self.log_viewer.add_log('INFO', 'Dashboard updated successfully')
            except (PermissionError, IOError, OSError) as e:
                error_msg = get_user_friendly_error(e)
                self.log_viewer.add_log('ERROR', f'Failed to write dashboard: {str(e)}')
                QMessageBox.warning(
                    self,
                    self.translations.get('warning_title'),
                    f"Failed to update dashboard:\n\n{error_msg}"
                )
            except Exception as e:
                error_msg = get_user_friendly_error(e)
                self.log_viewer.add_log('ERROR', f'Unexpected error writing dashboard: {str(e)}')
                QMessageBox.warning(
                    self,
                    self.translations.get('warning_title'),
                    f"Error updating dashboard:\n\n{error_msg}"
                )

            # Refresh archive
            self.refresh_archive()
            self.refresh_files()

            self.statusBar().showMessage(self.translations.get('processing_complete'))
            QMessageBox.information(
                self,
                self.translations.get('success_title'),
                self.translations.get('processing_complete')
            )

    def processing_error(self, error_msg: str):
        """
        Handle processing error.

        Re-enables the process button, hides progress bar, and shows user-friendly error message.

        Args:
            error_msg: Error message to display to user (should already be user-friendly)
        """
        self.process_btn.setEnabled(True)
        self.progress_bar.setVisible(False)

        # Error message should already be user-friendly from ProcessThread
        QMessageBox.critical(
            self,
            self.translations.get('error_title'),
            error_msg
        )

    def update_preview_table(self, summary_df: pd.DataFrame):
        """
        Update the preview table with category summary.

        Groups transactions by category, calculates totals and percentages,
        and displays in a color-coded table with grand total row.

        Args:
            summary_df: DataFrame with columns: year, month, category, subcat, monthly_amount
        """
        # Normalize data for consistent display/filtering
        summary_df = summary_df.copy()
        summary_df['category'] = summary_df['category'].astype(str)

        if summary_df.empty:
            self.preview_table.setRowCount(0)
            return

        # Update tab title with year
        years = sorted(summary_df['year'].unique())
        year_str = ", ".join(map(str, years))
        self.preview_tabs.setTabText(self.preview_tabs.indexOf(self.table_tab_widget),
                                     f"{self.translations.get('table_tab')} ({year_str})")

        # Group by Category and sum amounts
        category_summary = summary_df.groupby('category')['monthly_amount'].sum().reset_index()

        # Calculate grand total
        grand_total = category_summary['monthly_amount'].sum()

        # Calculate percentage for sorting
        category_summary['pct'] = (category_summary['monthly_amount'] / grand_total * 100) if grand_total > 0 else 0

        # Sort by percentage descending (highest first)
        category_summary = category_summary.sort_values('pct', ascending=False)

        logger = logging.getLogger(__name__)
        logger.debug("Sorted Category Summary:")
        logger.debug(f"\n{category_summary[['category', 'monthly_amount', 'pct']].head()}")

        # Store summary for filtering
        self.last_summary_df = summary_df

        # Store current filter state
        self.current_category_filter = None

        # Set up table
        self.preview_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.preview_table.setSortingEnabled(False)  # Ensure manual order is respected
        self.preview_table.setRowCount(len(category_summary) + 1)  # +1 for grand total
        self.preview_table.setColumnCount(3)
        self.preview_table.setAlternatingRowColors(True)

        # Enable row selection and add click handler
        self.preview_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.preview_table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)

        # Reconnect handlers to avoid duplicate connections after refresh
        try:
            self.preview_table.itemClicked.disconnect(self.on_category_row_clicked)
        except TypeError:
            pass
        try:
            self.preview_table.itemSelectionChanged.disconnect(self.on_category_selection_changed)
        except TypeError:
            pass

        self.preview_table.itemClicked.connect(self.on_category_row_clicked)
        self.preview_table.itemSelectionChanged.connect(self.on_category_selection_changed)

        headers = [
            self.translations.get('category'),
            self.translations.get('amount'),
            '%'  # Percentage of total
        ]
        self.preview_table.setHorizontalHeaderLabels(headers)

        # Category colors (cycling through pleasant colors)
        category_colors = [
            QColor(230, 240, 255),  # Light blue
            QColor(255, 240, 230),  # Light orange
            QColor(240, 255, 240),  # Light green
            QColor(255, 245, 230),  # Light yellow
            QColor(245, 230, 255),  # Light purple
            QColor(230, 255, 255),  # Light cyan
        ]

        for row_idx, (index, row) in enumerate(category_summary.iterrows()):
            cat_color = category_colors[row_idx % len(category_colors)]

            # Category
            cat_item = QTableWidgetItem(str(row['category']))
            cat_item.setFont(QFont('Arial', 10))
            cat_item.setBackground(cat_color)
            self.preview_table.setItem(row_idx, 0, cat_item)

            # Amount
            amount_item = QTableWidgetItem(f"₪{row['monthly_amount']:,.2f}")
            amount_item.setFont(QFont('Arial', 10))
            amount_item.setBackground(cat_color)
            self.preview_table.setItem(row_idx, 1, amount_item)

            # Percentage
            pct = (row['monthly_amount'] / grand_total * 100) if grand_total > 0 else 0
            pct_item = QTableWidgetItem(f"{pct:.1f}%")
            pct_item.setFont(QFont('Arial', 10))
            pct_item.setBackground(cat_color)
            self.preview_table.setItem(row_idx, 2, pct_item)

        # Add grand total row
        total_row = len(category_summary)

        total_label = QTableWidgetItem("TOTAL" if self.translations.language == 'en' else "סה\"כ כולל")
        total_label.setFont(QFont('Arial', 11, QFont.Weight.Bold))
        total_label.setBackground(QColor(180, 180, 180))
        self.preview_table.setItem(total_row, 0, total_label)

        total_value = QTableWidgetItem(f"₪{grand_total:,.2f}")
        total_value.setFont(QFont('Arial', 11, QFont.Weight.Bold))
        total_value.setBackground(QColor(180, 180, 180))
        self.preview_table.setItem(total_row, 1, total_value)

        total_pct = QTableWidgetItem("100.0%")
        total_pct.setFont(QFont('Arial', 11, QFont.Weight.Bold))
        total_pct.setBackground(QColor(180, 180, 180))
        self.preview_table.setItem(total_row, 2, total_pct)

        # Adjust column widths
        self.preview_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)  # Category
        self.preview_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)  # Amount
        self.preview_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)  # %

        # Show summary in status bar
        self.statusBar().showMessage(
            f"Total: ₪{grand_total:,.2f} | {len(category_summary)} categories"
        )

    def on_category_selection_changed(self):
        """Handle selection changes to drive chart updates."""
        selection_model = self.preview_table.selectionModel()
        if not selection_model:
            return

        selected_rows = selection_model.selectedRows()
        if not selected_rows:
            return

        self._handle_category_selection(selected_rows[0].row())

    def on_category_row_clicked(self, item: QTableWidgetItem):
        """
        Handle click on category table row to filter chart.

        Shows monthly breakdown bar chart for the selected category.

        Args:
            item: The table item that was clicked
        """
        self._handle_category_selection(item.row())

    def _handle_category_selection(self, row: int):
        """Shared handler for clicks and selection changes."""
        if not hasattr(self, 'last_summary_df') or self.last_summary_df is None:
            return
        if self.last_summary_df.empty:
            return

        # Don't filter on total row
        if row >= self.preview_table.rowCount() - 1:
            return

        category_item = self.preview_table.item(row, 0)
        if not category_item:
            return

        category = category_item.text()
        if category is None:
            return

        # Toggle filter: if same category clicked, show all; otherwise filter
        if self.current_category_filter == category:
            # Show all
            self.current_category_filter = None
            self.chart_widget.update_chart(
                self.last_summary_df,
                category_filter=None,
                full_summary_df=self.last_summary_df,
                show_all_callback=None
            )
            self.statusBar().showMessage(self.translations.get('showing_all_categories') or "Showing all categories")
            self.preview_tabs.setCurrentWidget(self.chart_widget)
        else:
            # Filter by category - show monthly breakdown
            self.current_category_filter = category
            filtered_df = self.last_summary_df[self.last_summary_df['category'] == category]
            self.chart_widget.update_chart(
                filtered_df,
                category_filter=category,
                full_summary_df=self.last_summary_df,
                show_all_callback=self.show_all_categories
            )
            self.statusBar().showMessage(
                self.translations.get('filtered_by_category', category=category) or f"Filtered by: {category}"
            )
            self.preview_tabs.setCurrentWidget(self.chart_widget)

    def show_all_categories(self, switch_tab: bool = False, clear_selection: bool = False):
        """Show all categories in chart (called from Show All/back buttons)."""
        self.current_category_filter = None
        if clear_selection:
            self.preview_table.clearSelection()
        self.chart_widget.update_chart(
            self.last_summary_df,
            category_filter=None,
            full_summary_df=self.last_summary_df,
            show_all_callback=None
        )
        if switch_tab:
            self.preview_tabs.setCurrentWidget(self.chart_widget)
        self.statusBar().showMessage(self.translations.get('showing_all_categories') or "Showing all categories")

    def clear_archive(self):
        """
        Clear all archived files.

        Prompts user for confirmation, then deletes all Excel files from the archive
        directory and refreshes the archive list.
        """
        reply = QMessageBox.question(
            self,
            self.translations.get('confirm_title'),
            self.translations.get('confirm_clear_archive'),
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )

        if reply == QMessageBox.StandardButton.Yes:
            archive_path = ARCHIVE_DIR
            if archive_path.exists():
                for file_path in archive_path.glob('*.xls*'):
                    file_path.unlink()

            self.refresh_archive()
            QMessageBox.information(
                self,
                self.translations.get('success_title'),
                self.translations.get('archive_cleared')
            )

    def show_category_management(self):
        """
        Show category management dialog.

        Opens a dialog allowing users to view, edit, and delete category mappings.
        """
        try:
            cat_mgr = CategoryManager(CATEGORIES_FILE_PATH, DASHBOARD_FILE_PATH)
            dialog = CategoryManagementDialog(cat_mgr, self.translations, self)
            dialog.exec()
        except Exception as e:
            error_msg = get_user_friendly_error(e)
            QMessageBox.warning(self, self.translations.get('warning_title'),
                              f"Failed to open category management:\n\n{error_msg}")

    def open_dashboard(self):
        """
        Open the dashboard Excel file in the default application.

        Uses the system's default application (typically Excel) to open the dashboard file.
        Shows warning if dashboard file is not found.
        """
        if DASHBOARD_FILE_PATH.exists():
            os.startfile(str(DASHBOARD_FILE_PATH))
            self.statusBar().showMessage(self.translations.get('dashboard_opened'))
        else:
            QMessageBox.warning(
                self,
                self.translations.get('warning_title'),
                self.translations.get('error_occurred', error='Dashboard file not found')
            )
def main():
    """
    Main entry point for GUI application.

    Creates QApplication, sets up global exception handler, creates and shows
    the main window, and starts the event loop.
    """
    app = QApplication(sys.argv)

    # Set application-wide font
    font = QFont('Arial', 10)
    app.setFont(font)

    # Global exception handler with error reporting
    from src.error_reporter import setup_error_reporting
    error_reporter = setup_error_reporting()

    def exception_hook(exctype, value, tb):
        error_id = error_reporter.capture_exception(exctype, value, tb)
        user_msg = error_reporter.generate_user_message(error_id, value)

        import traceback
        error_msg = ''.join(traceback.format_exception(exctype, value, tb))
        print(f"Unhandled exception (ID: {error_id}):\\n{error_msg}")

        QMessageBox.critical(None, "Fatal Error", user_msg)
        sys.exit(1)

    sys.excepthook = exception_hook

    window = BudgetTrackerGUI()
    window.show()

    sys.exit(app.exec())


if __name__ == '__main__':
    main()


