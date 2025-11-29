import openpyxl
import os
import shutil
import sys

def setup_dashboard():
    src_file = os.path.join('UserFiles', 'dashboard.xlsx')
    dst_file = os.path.join('deployment', 'UserFiles', 'dashboard.xlsx')

    print(f"Source: {src_file}")
    print(f"Destination: {dst_file}")

    if os.path.exists(src_file):
        print('Found source dashboard, copying...')
        try:
            shutil.copy2(src_file, dst_file)
            print('Copy successful. Cleaning data...')

            wb = openpyxl.load_workbook(dst_file)

            if 'Template' not in wb.sheetnames:
                print("Template sheet not found, creating it...")
                ws = wb.create_sheet('Template')
            else:
                ws = wb['Template']

            # Remove other sheets
            for sheet in list(wb.sheetnames):
                if sheet != 'Template':
                    print(f"Removing sheet: {sheet}")
                    del wb[sheet]

            # Clear data columns (3-14)
            print("Clearing data columns (skipping 'סיכום')...")
            for row in ws.iter_rows(min_row=2, max_col=14):
                # Check if this is the summary row
                first_col_val = str(row[0].value).strip() if row[0].value else ""
                second_col_val = str(row[1].value).strip() if row[1].value else ""

                if "סיכום" in first_col_val or "סיכום" in second_col_val:
                    print("Skipping summary row")
                    continue

                for cell in row[2:]:
                    cell.value = None

            wb.save(dst_file)
            print('Dashboard ready.')

        except Exception as e:
            print(f'ERROR processing dashboard: {e}')
            import traceback
            traceback.print_exc()
    else:
        print('Source dashboard not found. Creating new template...')
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Template"
            ws.append(['Category', 'Subcategory'] + [str(i) for i in range(1, 13)])
            wb.save(dst_file)
            print('New template created.')
        except Exception as e:
            print(f'ERROR creating new template: {e}')

if __name__ == "__main__":
    setup_dashboard()
