"""Test robustness against malformed category data."""
import pytest
import json
from src.category_manager import CategoryManager
from openpyxl import Workbook


def create_dashboard_with_template(file_path):
    """Helper to create a dashboard with Template sheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Template"
    ws['A1'] = 'Category'
    ws['B1'] = 'Subcategory'
    ws['A2'] = 'Shopping'
    ws['B2'] = 'Online'
    ws['A4'] = 'Food'
    ws['B4'] = 'Groceries'
    wb.save(file_path)


def test_category_manager_handles_empty_key(temp_dir):
    """Test that CategoryManager handles empty merchant key gracefully."""
    categories_file = temp_dir / 'categories.json'
    categories = {
        '': '',  # Invalid: empty key with string value
        'Amazon': ['Shopping', 'Online']
    }
    with open(categories_file, 'w', encoding='utf-8') as f:
        json.dump(categories, f)

    dashboard_file = temp_dir / 'dashboard.xlsx'
    create_dashboard_with_template(dashboard_file)

    manager = CategoryManager(categories_file, dashboard_file)
    
    # Should not crash on find_similar_merchant
    result = manager.find_similar_merchant('AMAZON')
    assert result == ('Shopping', 'Online')


def test_category_manager_handles_string_value(temp_dir):
    """Test that CategoryManager handles string value instead of list."""
    categories_file = temp_dir / 'categories.json'
    categories = {
        'BadMerchant': 'not_a_list',  # Invalid: string instead of list
        'GoodMerchant': ['Food', 'Groceries']
    }
    with open(categories_file, 'w', encoding='utf-8') as f:
        json.dump(categories, f)

    dashboard_file = temp_dir / 'dashboard.xlsx'
    create_dashboard_with_template(dashboard_file)

    manager = CategoryManager(categories_file, dashboard_file)
    
    # Should not crash when iterating
    result = manager.find_similar_merchant('GoodMerchant')
    assert result == ('Food', 'Groceries')
    
    # BadMerchant might match GoodMerchant via substring, which is OK
    # The important thing is it doesn't crash when iterating over bad entries
    result = manager.find_similar_merchant('BadMerchant')
    # Result might be ('Food', 'Groceries') due to substring match with GoodMerchant
    assert result is None or result == ('Food', 'Groceries')


def test_category_manager_handles_short_list(temp_dir):
    """Test that CategoryManager handles list with only 1 element."""
    categories_file = temp_dir / 'categories.json'
    categories = {
        'ShortList': ['OnlyCategory'],  # Invalid: only 1 element
        'GoodMerchant': ['Food', 'Groceries']
    }
    with open(categories_file, 'w', encoding='utf-8') as f:
        json.dump(categories, f)

    dashboard_file = temp_dir / 'dashboard.xlsx'
    create_dashboard_with_template(dashboard_file)

    manager = CategoryManager(categories_file, dashboard_file)
    
    # Should not crash
    result = manager.find_similar_merchant('GoodMerchant')
    assert result == ('Food', 'Groceries')
