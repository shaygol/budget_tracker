# ===== tests/test_category_manager.py =====
"""
Tests for the CategoryManager class.
"""
import pytest
import pandas as pd
import json
from pathlib import Path
from openpyxl import Workbook
from code.category_manager import CategoryManager


def create_dashboard_with_template(file_path):
    """Helper to create a dashboard with Template sheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Template"

    # Add headers
    ws['A1'] = 'Category'
    ws['B1'] = 'Subcategory'

    # Add some sample data to make template valid
    ws['A2'] = 'Shopping'
    ws['B2'] = 'Online'
    ws['A3'] = ''
    ws['B3'] = 'Retail'
    ws['A4'] = 'Food'
    ws['B4'] = 'Groceries'

    wb.save(file_path)


def test_category_manager_loads_categories(temp_dir):
    """Test that CategoryManager loads categories from JSON."""
    categories_file = temp_dir / 'categories.json'
    categories = {
        'Amazon': ['Shopping', 'Online'],
        'Supermarket': ['Food', 'Groceries']
    }
    with open(categories_file, 'w', encoding='utf-8') as f:
        json.dump(categories, f)

    dashboard_file = temp_dir / 'dashboard.xlsx'
    create_dashboard_with_template(dashboard_file)

    manager = CategoryManager(categories_file, dashboard_file)

    assert manager.category_map == categories


def test_category_manager_saves_categories(temp_dir):
    """Test that CategoryManager saves new categories."""
    categories_file = temp_dir / 'categories.json'
    categories_file.write_text('{}', encoding='utf-8')

    dashboard_file = temp_dir / 'dashboard.xlsx'
    create_dashboard_with_template(dashboard_file)

    manager = CategoryManager(categories_file, dashboard_file)
    manager.category_map['NewMerchant'] = ['Category', 'Subcategory']
    manager.save_categories()

    # Read back
    with open(categories_file, 'r', encoding='utf-8') as f:
        saved = json.load(f)

    assert 'NewMerchant' in saved
    assert saved['NewMerchant'] == ['Category', 'Subcategory']


def test_category_manager_maps_known_merchants(temp_dir, sample_categories):
    """Test that known merchants are mapped correctly."""
    categories_file = temp_dir / 'categories.json'
    with open(categories_file, 'w', encoding='utf-8') as f:
        json.dump(sample_categories, f)

    dashboard_file = temp_dir / 'dashboard.xlsx'
    create_dashboard_with_template(dashboard_file)

    df = pd.DataFrame({
        'merchant': ['Amazon', 'Supermarket'],
        'amount': [100.0, 50.0]
    })

    manager = CategoryManager(categories_file, dashboard_file)
    df['category'] = df['merchant'].map(lambda m: manager.category_map.get(m, [None, None])[0])
    df['subcat'] = df['merchant'].map(lambda m: manager.category_map.get(m, [None, None])[1])

    assert df['category'].tolist() == ['Shopping', 'Food']
    assert df['subcat'].tolist() == ['Online', 'Groceries']
