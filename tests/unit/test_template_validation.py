"""Unit tests for template validation in category_manager.py."""

from pathlib import Path
from openpyxl import Workbook
from src.category_manager import (
    CategoryManager,
    ValidationResult,
    is_header_value,
    normalize_category_name,
    safe_get_cell_value,
)


def _create_test_dashboard(path: Path, data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Template"
    ws["A1"] = "נושא הוצאה"
    ws["B1"] = "פירוט הוצאות"
    for i, (cat, subcat) in enumerate(data, start=2):
        ws[f"A{i}"] = cat
        ws[f"B{i}"] = subcat
    wb.save(path)


def test_template_helper_functions_and_validation_result():
    wb = Workbook()
    ws = wb.active

    safe_cases = [
        ("Test Category", "Test Category"),
        (123, "123"),
        (None, None),
        ("   ", None),
        ("#REF!", None),
        ("  Test  ", "Test"),
    ]
    for raw, expected in safe_cases:
        ws["A1"] = raw
        assert safe_get_cell_value(ws["A1"]) == expected

    normalize_cases = [
        ("הוצאות בית", "הוצאות בית"),
        ("  קניות  ", "קניות"),
        ("קניות    חודשיות", "קניות חודשיות"),
        (None, None),
        ("", None),
        ("   ", None),
    ]
    for raw, expected in normalize_cases:
        assert normalize_category_name(raw) == expected

    header_true_cases = [
        ("נושא", "category"),
        ("נושא הוצאה", "category"),
        ("Category", "category"),
        ("category", "category"),
        ("פירוט", "subcategory"),
        ("פירוט הוצאות", "subcategory"),
    ]
    for value, header_type in header_true_cases:
        assert is_header_value(value, header_type) is True

    header_false_cases = [
        ("קניות", "category"),
        ("מזון", "subcategory"),
        (None, "category"),
    ]
    for value, header_type in header_false_cases:
        assert is_header_value(value, header_type) is False

    result = ValidationResult()
    assert result.is_valid is True
    assert result.errors == []
    assert result.warnings == []

    result.add_warning("warn")
    assert result.is_valid is True
    assert result.warnings == ["warn"]

    result.add_error("err")
    assert result.is_valid is False
    assert result.errors == ["err"]


def test_template_validation_scenarios(tmp_path):
    categories_path = tmp_path / "categories.json"
    categories_path.write_text("{}")

    long_name = "א" * 150
    scenarios = [
        {
            "name": "valid",
            "data": [("הוצאות בית", "חשמל"), ("", "גז"), ("קניות", "מזון"), ("", "ביגוד")],
            "expect_valid": True,
            "error_contains": None,
            "warning_contains": None,
        },
        {
            "name": "dup_category",
            "data": [("הוצאות בית", "חשמל"), ("קניות", "מזון"), ("הוצאות בית", "מים")],
            "expect_valid": False,
            "error_contains": "Duplicate category",
            "warning_contains": None,
        },
        {
            "name": "dup_subcategory",
            "data": [("הוצאות בית", "חשמל"), ("", "גז"), ("", "חשמל")],
            "expect_valid": False,
            "error_contains": "Duplicate subcategory",
            "warning_contains": None,
        },
        {
            "name": "empty",
            "data": [],
            "expect_valid": False,
            "error_contains": "empty",
            "warning_contains": None,
        },
        {
            "name": "no_subcat",
            "data": [("הוצאות בית", None)],
            "expect_valid": True,
            "error_contains": None,
            "warning_contains": "no subcategories",
        },
        {
            "name": "single_subcat",
            "data": [("הוצאות בית", "חשמל")],
            "expect_valid": True,
            "error_contains": None,
            "warning_contains": "only one subcategory",
        },
        {
            "name": "long_name",
            "data": [(long_name, "test")],
            "expect_valid": True,
            "error_contains": None,
            "warning_contains": "very long",
        },
        {
            "name": "suspicious_chars",
            "data": [("קניות|מזון", "test"), ("", "test2;data")],
            "expect_valid": True,
            "error_contains": None,
            "warning_contains": "suspicious characters",
        },
    ]

    for scenario in scenarios:
        dashboard_path = tmp_path / f"dashboard_{scenario['name']}.xlsx"
        _create_test_dashboard(dashboard_path, scenario["data"])
        manager = CategoryManager(categories_path, dashboard_path, strict_validation=False)
        validation = manager.validate_template_structure()

        assert validation.is_valid is scenario["expect_valid"], scenario["name"]

        if scenario["error_contains"]:
            assert any(
                scenario["error_contains"].lower() in err.lower()
                for err in validation.errors
            ), scenario["name"]
        else:
            assert len(validation.errors) == 0 or scenario["expect_valid"] is False

        if scenario["warning_contains"]:
            assert any(
                scenario["warning_contains"].lower() in warn.lower()
                for warn in validation.warnings
            ), scenario["name"]
