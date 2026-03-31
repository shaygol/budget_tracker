"""
Tests for GUI threading components.
"""
import pytest
import pandas as pd
from PyQt5.QtWidgets import QApplication
from PyQt5.QtCore import QTimer
from gui_app import ProcessThread
from src.translations import Translations
from src.config import TRANSACTIONS_DIR
from pathlib import Path





@pytest.fixture
def translations():
    """Create translations object."""
    return Translations('en')


class TestProcessThread:
    """Tests for ProcessThread."""

    def test_thread_creation(self, qapp, translations):
        """Test that thread can be created."""
        thread = ProcessThread(translations)
        assert thread is not None
        assert thread.translations == translations
        assert thread.category_response is None
        assert thread.response_ready is False
        assert thread._should_stop is False

    def test_thread_stop(self, qapp, translations):
        """Test stopping the thread."""
        thread = ProcessThread(translations)
        thread.stop()

        assert thread._should_stop is True

    def test_timeout_handling(self, qapp, translations):
        """Test timeout timer setup."""
        thread = ProcessThread(translations)

        assert thread.timeout_timer is not None
        assert thread.timeout_timer.isSingleShot() is True

    def test_map_categories_gui_no_unknown(self, qapp, translations, tmp_path):
        """Test category mapping with no unknown merchants."""
        from src.category_manager import CategoryManager
        from openpyxl import Workbook

        # Create test dashboard
        dashboard_file = tmp_path / 'dashboard.xlsx'
        wb = Workbook()
        ws = wb.active
        ws.title = 'Template'
        ws['A1'] = 'Category'
        ws['B1'] = 'Subcategory'
        ws['A2'] = 'Food'
        ws['B2'] = 'Groceries'
        wb.save(dashboard_file)

        # Create categories file
        categories_file = tmp_path / 'categories.json'
        import json
        with open(categories_file, 'w') as f:
            json.dump({'Merchant1': ['Food', 'Groceries']}, f)

        thread = ProcessThread(translations)
        cat_mgr = CategoryManager(categories_file, dashboard_file)

        df = pd.DataFrame({
            'merchant': ['Merchant1', 'Merchant1'],
            'amount': [100.0, 200.0]
        })

        result_df = thread._map_categories_gui(df, cat_mgr)

        assert 'category' in result_df.columns
        assert 'subcat' in result_df.columns
        assert result_df['category'].iloc[0] == 'Food'
        assert result_df['subcat'].iloc[0] == 'Groceries'

    def test_map_categories_gui_prompts_once_for_default_mapped_merchant(self, qapp, translations, tmp_path, monkeypatch):
        """Default mapping should prompt once, then persist and stop prompting."""
        from src.category_manager import CategoryManager
        from openpyxl import Workbook

        dashboard_file = tmp_path / 'dashboard.xlsx'
        wb = Workbook()
        ws = wb.active
        ws.title = 'Template'
        ws['A1'] = 'Category'
        ws['B1'] = 'Subcategory'
        ws['A2'] = 'Food'
        ws['B2'] = 'Groceries'
        wb.save(dashboard_file)

        categories_file = tmp_path / 'categories.json'
        import json
        with open(categories_file, 'w') as f:
            json.dump({}, f)

        df = pd.DataFrame({
            'merchant': ['Merchant1'],
            'amount': [100.0]
        })

        # First run: prompt is expected for default-only mapping.
        thread1 = ProcessThread(translations)
        cat_mgr1 = CategoryManager(categories_file, dashboard_file)
        cat_mgr1.category_map = {'Merchant1': ['Food', 'Groceries']}
        monkeypatch.setattr(cat_mgr1, '_load_default_categories', lambda: {'Merchant1': ['Food', 'Groceries']})

        prompted_first = []
        def _first_prompt(*_):
            prompted_first.append(True)
            thread1.category_response = ('Food', 'Groceries')
            thread1.response_ready = True
        thread1.category_needed.connect(_first_prompt)

        result_df_1 = thread1._map_categories_gui(df, cat_mgr1)
        assert len(prompted_first) == 1
        assert result_df_1['category'].iloc[0] == 'Food'
        assert result_df_1['subcat'].iloc[0] == 'Groceries'

        # Second run: should not prompt again (confirmation persisted in user categories).
        thread2 = ProcessThread(translations)
        cat_mgr2 = CategoryManager(categories_file, dashboard_file)
        cat_mgr2.category_map['Merchant1'] = ['Food', 'Groceries']
        monkeypatch.setattr(cat_mgr2, '_load_default_categories', lambda: {'Merchant1': ['Food', 'Groceries']})

        prompted_second = []
        def _second_prompt(*_):
            prompted_second.append(True)
            thread2.category_response = ('Food', 'Groceries')
            thread2.response_ready = True
        thread2.category_needed.connect(_second_prompt)

        result_df_2 = thread2._map_categories_gui(df, cat_mgr2)
        assert len(prompted_second) == 0
        assert result_df_2['category'].iloc[0] == 'Food'
        assert result_df_2['subcat'].iloc[0] == 'Groceries'

