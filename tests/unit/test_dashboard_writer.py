# ===== tests/test_dashboard_writer.py =====
"""
Tests for the DashboardWriter class.
"""
import pytest
import pandas as pd
from pathlib import Path
from openpyxl import Workbook, load_workbook
from src.dashboard_writer import DashboardWriter


def _build_dashboard_with_summary() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "2024"
    ws["A1"] = "Category"
    ws["B1"] = "Subcategory"
    ws["C1"] = 1
    ws["A2"] = "Food"
    ws["B2"] = "Groceries"
    ws["A3"] = "Bank"
    ws["B3"] = "Fees"
    ws["A4"] = "Summary"
    ws.merge_cells("A4:B4")
    ws["C4"] = "=SUM(C2:C3)"
    return wb


def _build_template_dashboard_with_summary() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Template"
    ws["A1"] = "Category"
    ws["B1"] = "Subcategory"
    ws["C1"] = 1
    ws["A2"] = "Food"
    ws["B2"] = "Groceries"
    ws["A3"] = "Bank"
    ws["B3"] = "Fees"
    ws["A4"] = "Summary"
    ws.merge_cells("A4:B4")
    ws["C4"] = "=SUM(C2:C3)"
    return wb


def test_dashboard_writer_validates_summary():
    """Test that DashboardWriter validates summary DataFrame."""
    writer = DashboardWriter(Path('dummy.xlsx'))

    # Missing columns
    invalid_df = pd.DataFrame({'year': [2024]})
    assert writer._validate_summary(invalid_df) is False

    # Valid columns
    valid_df = pd.DataFrame({
        'year': [2024],
        'month': [1],
        'category': ['Food'],
        'subcat': ['Groceries'],
        'monthly_amount': [100.0]
    })
    assert writer._validate_summary(valid_df) is True


def test_dashboard_writer_validates_empty_dataframe():
    """Test that DashboardWriter rejects empty DataFrames."""
    writer = DashboardWriter(Path('dummy.xlsx'))

    empty_df = pd.DataFrame(columns=['year', 'month', 'category', 'subcat', 'monthly_amount'])
    assert writer._validate_summary(empty_df) is False


def test_dashboard_writer_conflict_resolver_callback():
    """Test that conflict resolver callback is used when provided."""
    writer = DashboardWriter(Path('dummy.xlsx'))

    # Mock conflict resolver
    def resolver(month_key):
        return 'override'

    result = writer._prompt_user_decision('2024-1', conflict_resolver=resolver)
    assert result == 'override'


def test_dashboard_writer_gets_category_ranges(temp_dir):
    """Test that category row ranges are correctly identified."""
    # Create a simple workbook
    wb = Workbook()
    ws = wb.active
    ws['A2'] = 'Food'
    ws['A3'] = 'Food'
    ws['A4'] = 'Shopping'

    dashboard_path = temp_dir / 'test_dashboard.xlsx'
    wb.save(dashboard_path)

    writer = DashboardWriter(dashboard_path)
    ranges = writer._get_category_row_ranges(ws)

    assert 'Food' in ranges
    assert 'Shopping' in ranges


def test_dashboard_writer_excludes_summary_row_from_category_ranges(temp_dir):
    """Summary row should not be treated as a category bucket."""
    wb = _build_dashboard_with_summary()
    dashboard_path = temp_dir / 'test_dashboard.xlsx'
    wb.save(dashboard_path)

    writer = DashboardWriter(dashboard_path)
    ranges = writer._get_category_row_ranges(wb["2024"])

    assert "Summary" not in ranges
    assert ranges["Food"] == (2, 2)
    assert ranges["Bank"] == (3, 3)


def test_dashboard_writer_adds_new_category_before_summary_row(temp_dir):
    """New categories should be inserted before the merged summary row."""
    wb = _build_dashboard_with_summary()
    dashboard_path = temp_dir / 'test_dashboard.xlsx'
    wb.save(dashboard_path)

    writer = DashboardWriter(dashboard_path)
    ws = wb["2024"]

    writer._add_new_category(ws, "Shopping")

    assert ws["A4"].value == "Shopping"
    assert ws["A5"].value == "Summary"
    assert "A5:B5" in {str(rng) for rng in ws.merged_cells.ranges}


def test_dashboard_writer_adds_new_subcategory_before_summary_row(temp_dir):
    """New subcategories should shift the summary row instead of writing into merged cells."""
    wb = _build_dashboard_with_summary()
    dashboard_path = temp_dir / 'test_dashboard.xlsx'
    wb.save(dashboard_path)

    writer = DashboardWriter(dashboard_path)
    ws = wb["2024"]
    ranges = writer._get_category_row_ranges(ws)
    subcat_map = writer._build_subcat_location_map(ws, ranges)

    writer._add_new_subcategory(ws, "Bank", "Interest", subcat_map, ranges)

    assert ws["A3"].value == "Bank"
    assert ws["B4"].value == "Interest"
    assert ws["A5"].value == "Summary"
    assert "A3:A4" in {str(rng) for rng in ws.merged_cells.ranges}
    assert "A5:B5" in {str(rng) for rng in ws.merged_cells.ranges}


def test_dashboard_writer_update_adds_new_category_before_summary_row(temp_dir, monkeypatch):
    """Public update() should insert new categories safely before merged summary rows."""
    backup_dir = temp_dir / "dash_backups"
    backup_dir.mkdir()
    monkeypatch.setattr("src.dashboard_writer.DASHBOARD_BACKUP_DIR", backup_dir)

    dashboard_path = temp_dir / "dashboard.xlsx"
    _build_template_dashboard_with_summary().save(dashboard_path)

    summary_df = pd.DataFrame({
        "year": [2024],
        "month": [1],
        "category": ["Travel"],
        "subcat": ["Flights"],
        "monthly_amount": [123.45],
    })

    writer = DashboardWriter(dashboard_path)
    writer.update(summary_df, conflict_resolver=lambda _: "override")

    wb = load_workbook(dashboard_path)
    ws = wb["2024"]

    assert ws["A6"].value == "Summary"
    assert "A4:A5" in {str(rng) for rng in ws.merged_cells.ranges}
    assert "A6:B6" in {str(rng) for rng in ws.merged_cells.ranges}
    assert ws["B5"].value == "Flights"
    assert ws["C5"].value == 123.45
    wb.close()


def test_dashboard_writer_update_adds_new_subcategory_before_summary_row(temp_dir, monkeypatch):
    """Public update() should shift merged summary rows when inserting subcategories."""
    backup_dir = temp_dir / "dash_backups"
    backup_dir.mkdir()
    monkeypatch.setattr("src.dashboard_writer.DASHBOARD_BACKUP_DIR", backup_dir)

    dashboard_path = temp_dir / "dashboard.xlsx"
    _build_template_dashboard_with_summary().save(dashboard_path)

    summary_df = pd.DataFrame({
        "year": [2024],
        "month": [1],
        "category": ["Food"],
        "subcat": ["Dining"],
        "monthly_amount": [50.0],
    })

    writer = DashboardWriter(dashboard_path)
    writer.update(summary_df, conflict_resolver=lambda _: "override")

    wb = load_workbook(dashboard_path)
    ws = wb["2024"]

    assert ws["A5"].value == "Summary"
    assert "A2:A3" in {str(rng) for rng in ws.merged_cells.ranges}
    assert "A5:B5" in {str(rng) for rng in ws.merged_cells.ranges}
    assert ws["B3"].value == "Dining"
    assert ws["C3"].value == 50.0
    wb.close()
