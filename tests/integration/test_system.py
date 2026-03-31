import pytest
import pandas as pd
import os
from pathlib import Path
from unittest.mock import patch
from src.file_manager import load_transaction_files
from src.normalizer import Normalizer
from src.category_manager import CategoryManager
from src.dashboard_writer import DashboardWriter
from src.previewer import Previewer
from openpyxl import Workbook

class TestSystemWorkflow:
    @pytest.fixture
    def setup_env(self, tmp_path):
        # Create structure
        user_files = tmp_path / "UserFiles"
        user_files.mkdir()

        # Create backups directory (which is now the transactions dir)
        backups = user_files / "backups"
        backups.mkdir()

        temp_proc = backups  # Transactions are now processed from backups dir

        archive = backups / "archive"
        archive.mkdir()

        dashboard_backups = backups / "dashboard"
        dashboard_backups.mkdir()

        # Create dummy categories.json
        cats_file = user_files / "categories.json"
        cats_file.write_text('{"Food": ["Groceries"]}', encoding='utf-8')

        # Create dummy dashboard.xlsx with Template sheet
        dash_file = user_files / "dashboard.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Template"
        ws["A1"] = "Category"
        ws["B1"] = "Subcategory"
        for month in range(1, 13):
            ws.cell(row=1, column=month + 2, value=str(month))
        ws["A2"] = "Food"
        ws["B2"] = "Groceries"
        ws["A3"] = "Summary"
        ws.merge_cells("A3:B3")
        ws["C3"] = "=SUM(C2:C2)"
        wb.save(dash_file)

        return {
            "root": user_files,
            "temp": temp_proc,
            "archive": archive,
            "dash_backups": dashboard_backups,
            "dash_file": dash_file,
            "cats_file": cats_file
        }

    def test_full_workflow(self, setup_env):
        env = setup_env

        # Create a transaction file
        trans_file = env["temp"] / "test_trans.xlsx"
        # Create valid content with enough columns to pass header detection
        # Logic requires > 3 matching columns (date, merchant, amount + 1 optional)
        df = pd.DataFrame({
            'תאריך': ['01/01/2025'],
            'שם בית העסק': ['Supermarket'],
            'סכום': [100.0],
            'שם כרטיס': ['1234']  # Added optional column
        })
        df.to_excel(trans_file, index=False)

        # Patch paths in the modules where they are used
        with patch('src.file_manager.ARCHIVE_DIR', env["archive"]), \
             patch('src.file_manager.TRANSACTIONS_DIR', env["temp"]), \
             patch('src.dashboard_writer.DASHBOARD_BACKUP_DIR', env["dash_backups"]):

            # 1. Load
            dfs = load_transaction_files(env["temp"])
            assert len(dfs) == 1
            assert trans_file.exists()

            # 2. Normalize
            normalizer = Normalizer()
            normalized_df = normalizer.normalize(dfs[0])
            assert not normalized_df.empty

            # 3. Categorize
            cat_mgr = CategoryManager(env["cats_file"], env["dash_file"])
            # Pre-seed category map to avoid user input
            cat_mgr.category_map['Supermarket'] = ['Food', 'Groceries']

            categorized_df = cat_mgr.map_categories(normalized_df)
            assert 'category' in categorized_df.columns

            # 4. Preview/Summary
            previewer = Previewer()
            summary = previewer.preview(categorized_df, confirm=False)
            assert not summary.empty

            # 5. Write to Dashboard
            writer = DashboardWriter(env["dash_file"])
            writer.update(summary)

            # Verify Dashboard Backup created
            backups = list(env["dash_backups"].glob("dashboard_*.xlsx"))
            assert len(backups) == 1

            # Verify Dashboard updated (check if 2025 sheet exists)
            import openpyxl
            wb = openpyxl.load_workbook(env["dash_file"])
            assert "2025" in wb.sheetnames
            wb.close()

    def test_full_workflow_with_new_category_on_merged_template(self, setup_env):
        env = setup_env

        trans_file = env["temp"] / "test_trans.xlsx"
        df = pd.DataFrame({
            'תאריך': ['01/01/2025'],
            'שם בית העסק': ['Supermarket'],
            'סכום': [100.0],
            'שם כרטיס': ['1234']
        })
        df.to_excel(trans_file, index=False)

        with patch('src.file_manager.ARCHIVE_DIR', env["archive"]), \
             patch('src.file_manager.TRANSACTIONS_DIR', env["temp"]), \
             patch('src.dashboard_writer.DASHBOARD_BACKUP_DIR', env["dash_backups"]):

            dfs = load_transaction_files(env["temp"])
            normalizer = Normalizer()
            normalized_df = normalizer.normalize(dfs[0])
            categorized_df = normalized_df.copy()
            categorized_df['category'] = 'Travel'
            categorized_df['subcat'] = 'Flights'
            previewer = Previewer()
            summary = previewer.preview(categorized_df, confirm=False)

            writer = DashboardWriter(env["dash_file"])
            writer.update(summary, conflict_resolver=lambda _: 'override')

            import openpyxl
            wb = openpyxl.load_workbook(env["dash_file"])
            ws = wb["2025"]
            merged_ranges = {str(rng) for rng in ws.merged_cells.ranges}

            assert ws["A5"].value == "Summary"
            assert "A3:A4" in merged_ranges
            assert "A5:B5" in merged_ranges
            assert ws["B4"].value == "Flights"
            assert ws["C4"].value == 100.0
            wb.close()
