import pandas as pd
import os
import logging
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet
from code.config import TEMPLATE_SHEET_NAME

logger = logging.getLogger(__name__)


class DashboardWriter:
    def __init__(self, dashboard_path):
        self.dashboard_path = dashboard_path
        self.template_sheet_name = TEMPLATE_SHEET_NAME
        self.user_decisions = {}

    def update(self, summary_df: pd.DataFrame, conflict_resolver=None):
        if not self._validate_summary(summary_df):
            return
            
        self.conflict_resolver = conflict_resolver

        years = sorted(summary_df['year'].unique())
        file_exists = os.path.exists(self.dashboard_path)

        if not file_exists:
            logger.error(f"Dashboard file not found: {self.dashboard_path}")
            return

        # Create backup before modifying dashboard
        import shutil
        from datetime import datetime
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            backup_path = f"{self.dashboard_path}.{timestamp}.backup"
            shutil.copy2(self.dashboard_path, backup_path)
            logger.info(f"Created backup: {backup_path}")
        except Exception as e:
            logger.warning(f"Failed to create backup: {e}")

        wb = load_workbook(self.dashboard_path)
        if self.template_sheet_name not in wb.sheetnames:
            logger.error(f"Template sheet '{self.template_sheet_name}' not found in dashboard.")
            return

        # First pass: create sheets per year if missing
        for year in years:
            sheet_name = str(year)
            if sheet_name not in wb.sheetnames:
                self._clone_template_sheet(wb, sheet_name)

        wb.save(self.dashboard_path)
        wb.close()

        # Second pass: update each year's sheet
        for year in years:
            self._update_sheet(year, summary_df[summary_df['year'] == year])

    def _validate_summary(self, df: pd.DataFrame) -> bool:
        required_cols = {'year', 'month', 'category', 'subcat', 'monthly_amount'}
        if not required_cols.issubset(df.columns):
            logger.error(f"Missing required columns: {required_cols - set(df.columns)}")
            return False
        if df.empty:
            logger.warning("Summary DataFrame is empty. Nothing will be written to the dashboard.")
            return False
        return True

    def _clone_template_sheet(self, wb, new_sheet_name: str):
        source = wb[self.template_sheet_name]
        target = wb.copy_worksheet(source)
        target.title = new_sheet_name
        target.sheet_view.rightToLeft = source.sheet_view.rightToLeft
        logger.info(f"Created new sheet by cloning template: {new_sheet_name}")

    def _update_sheet(self, year: int, df: pd.DataFrame):
        wb = load_workbook(self.dashboard_path)
        sheet_name = str(year)
        ws = wb[sheet_name]

        # Map month numbers to column indexes (e.g., Jan -> col 3)
        month_columns = {str(m): 2 + (m - 1) for m in range(1, 13)}

        # Build mapping of existing subcategories to rows
        category_ranges = self._get_category_row_ranges(ws)
        existing_map = self._build_subcat_location_map(ws, category_ranges)

        for _, row in df.iterrows():
            cat = row['category']
            subcat = row['subcat']
            month = str(int(row['month']))
            amount = row['monthly_amount']
            col = month_columns[month]

            if cat not in category_ranges:
                logger.info(f"New category detected: {cat}. Adding it.")
                self._add_new_category(ws, cat)
                category_ranges = self._get_category_row_ranges(ws)

            if (cat, subcat) not in existing_map:
                logger.info(f"New subcategory '{subcat}' under '{cat}' detected. Adding it.")
                self._add_new_subcategory(ws, cat, subcat, existing_map, category_ranges)
                existing_map = self._build_subcat_location_map(ws, category_ranges)

            row_idx = existing_map[(cat, subcat)]
            cell = ws.cell(row=row_idx, column=col + 1)
            existing_value = cell.value

            if existing_value is not None and existing_value != 0:
                month_key = f"{year}-{month}"
                if month_key not in self.user_decisions:
                    decision = self._prompt_user_decision(month_key, self.conflict_resolver)
                    self.user_decisions[month_key] = decision
                else:
                    decision = self.user_decisions[month_key]

                if decision == "override":
                    cell.value = amount
                elif decision == "add":
                    try:
                        cell.value = float(existing_value) + amount
                    except Exception:
                        cell.value = amount
                elif decision == "skip":
                    continue
            else:
                # No existing value, safe to write
                cell.value = amount

            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=False)

        wb.save(self.dashboard_path)
        wb.close()
        logger.info(f"Dashboard sheet updated for year {year}")


    def _prompt_user_decision(self, month_key: str, conflict_resolver=None) -> str:
        # Use callback if provided (GUI mode), otherwise use input (CLI mode)
        if conflict_resolver:
            return conflict_resolver(month_key)
            
        print(f"\nData already exists for {month_key}. Choose how to handle it:")
        print("1. Override existing data")
        print("2. Add to existing data")
        print("3. Skip this month")
        choice = input("Enter 1/2/3: ").strip()
        return {"1": "override", "2": "add", "3": "skip"}.get(choice, "skip")

    def _get_category_row_ranges(self, ws: Worksheet):
        # Returns a map: category name -> (start_row, end_row)
        ranges = {}
        current_cat = None
        start = None
        for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            cat_cell = row[0]
            if cat_cell.value:
                if current_cat and start:
                    ranges[current_cat] = (start, idx - 1)
                current_cat = cat_cell.value
                start = idx
        if current_cat and start:
            ranges[current_cat] = (start, idx)
        return ranges

    def _build_subcat_location_map(self, ws: Worksheet, cat_ranges: dict):
        # Returns a map: (category, subcategory) -> row index
        mapping = {}
        for cat, (start, end) in cat_ranges.items():
            for r in range(start, end + 1):
                subcat_cell = ws.cell(row=r, column=2)
                val = subcat_cell.value
                if isinstance(val, str):
                    mapping[(cat, val.strip())] = r
        return mapping

    def _add_new_category(self, ws: Worksheet, category: str):
        # Adds a new row for a new category at the bottom
        last_row = ws.max_row + 1
        ws.insert_rows(last_row)
        ws.cell(row=last_row, column=1, value=category)
        for col in range(2, 15):
            ws.cell(row=last_row, column=col, value="")

    def _add_new_subcategory(self, ws: Worksheet, category: str, subcat: str,
                             subcat_map: dict, cat_ranges: dict):
        # Inserts a subcategory row below the existing category group
        start, end = cat_ranges[category]
        insert_at = end + 1
        ws.insert_rows(insert_at)
        ws.cell(row=insert_at, column=2, value=subcat)
        for col in range(3, 15):
            ws.cell(row=insert_at, column=col, value="")
        # Re-merge the category header cell to extend one more row
        ws.unmerge_cells(start_row=start, end_row=end, start_column=1, end_column=1)
        ws.merge_cells(start_row=start, end_row=end + 1, start_column=1, end_column=1)
        cat_ranges[category] = (start, end + 1)
