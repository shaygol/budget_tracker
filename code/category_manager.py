import json
import sys
import logging
import re
from typing import List, Tuple, Dict, Optional
from dataclasses import dataclass, field
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.cell import Cell
from code.previewer import format_prompt
from code.config import TEMPLATE_SHEET_NAME, MAX_CATEGORIES

logger = logging.getLogger(__name__)

# Template validation constants
MAX_SUBCATEGORIES_PER_CATEGORY = 100
MAX_CATEGORY_NAME_LENGTH = 100
HEADER_PATTERNS = {
    'category': ['נושא', 'נושא הוצאה', 'category', 'קטגוריה'],
    'subcategory': ['פירוט', 'פירוט הוצאות', 'subcategory', 'תת-קטגוריה'],
}


@dataclass
class ValidationResult:
    """Results from template validation."""
    is_valid: bool = True
    errors: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)

    def add_error(self, message: str) -> None:
        """Add an error message."""
        self.errors.append(message)
        self.is_valid = False

    def add_warning(self, message: str) -> None:
        """Add a warning message."""
        self.warnings.append(message)


def safe_get_cell_value(cell: Cell) -> Optional[str]:
    """
    Safely extract cell value, handling errors and formulas.

    Args:
        cell: Excel cell to read

    Returns:
        Cleaned string value or None if cell is empty/invalid
    """
    val = cell.value

    # Handle None
    if val is None:
        return None

    # Handle Excel errors (e.g., #REF!, #VALUE!)
    if isinstance(val, str) and val.startswith('#'):
        logger.warning(f"Excel error in cell {cell.coordinate}: {val}")
        return None

    # Convert to string and strip whitespace
    val_str = str(val).strip()

    # Return None for empty strings
    if not val_str:
        return None

    return val_str


def normalize_category_name(name: Optional[str]) -> Optional[str]:
    """
    Normalize category name for consistent matching.

    Args:
        name: Category or subcategory name

    Returns:
        Normalized name or None if invalid
    """
    if not name:
        return None

    # Strip whitespace
    name = str(name).strip()

    # Return None for empty
    if not name:
        return None

    # Collapse multiple spaces
    name = ' '.join(name.split())

    return name


def is_header_value(value: Optional[str], header_type: str) -> bool:
    """
    Check if a value looks like a header.

    Args:
        value: Cell value to check
        header_type: Type of header ('category' or 'subcategory')

    Returns:
        True if value matches header patterns
    """
    if not value:
        return False

    value_lower = value.lower().strip()
    patterns = HEADER_PATTERNS.get(header_type, [])

    return any(pattern.lower() in value_lower for pattern in patterns)


class CategoryManager:
    def __init__(self, categories_path: str | Path, dashboard_path: str | Path, strict_validation: bool = True) -> None:
        self.categories_path = categories_path
        self.dashboard_path = dashboard_path
        self.strict_validation = strict_validation

        self.category_map = self._load_category_map()
        self.valid_categories = self.load_category_structure_from_template(strict=strict_validation)

    def _load_category_map(self) -> Dict[str, List[str]]:
        try:
            with open(self.categories_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except (FileNotFoundError, json.JSONDecodeError):
            logger.warning(f"Categories file not found or invalid: {self.categories_path}. Starting with empty map.")
            return {}

    def validate_template_structure(self) -> ValidationResult:
        """
        Validate the template structure for data quality issues.

        Returns:
            ValidationResult with errors and warnings
        """
        result = ValidationResult()

        # Check if file exists
        if not Path(self.dashboard_path).exists():
            result.add_error(f"Dashboard file not found: {self.dashboard_path}")
            return result

        wb = load_workbook(self.dashboard_path, data_only=True)
        try:
            # Check if Template sheet exists
            if TEMPLATE_SHEET_NAME not in wb.sheetnames:
                result.add_error(
                    f"Template sheet '{TEMPLATE_SHEET_NAME}' not found. "
                    f"Available sheets: {', '.join(wb.sheetnames)}"
                )
                return result

            ws = wb[TEMPLATE_SHEET_NAME]

            # Track for duplicate detection
            seen_categories = set()
            category_subcats = {}  # Track subcats per category for duplicate detection

            current_category = None
            row_num = 0

            for row in ws.iter_rows(min_row=2, max_col=2):
                row_num += 1
                cat_cell, subcat_cell = row

                # Safe cell reading
                cat_val = safe_get_cell_value(cat_cell)
                subcat_val = safe_get_cell_value(subcat_cell)

                # Skip header rows
                if is_header_value(cat_val, 'category') or is_header_value(subcat_val, 'subcategory'):
                    continue

                # Process category
                if cat_val:
                    cat_normalized = normalize_category_name(cat_val)

                    if cat_normalized:
                        # Check for duplicate categories
                        if cat_normalized in seen_categories:
                            result.add_error(
                                f"Duplicate category '{cat_normalized}' found at row {row_num + 1}"
                            )
                        else:
                            seen_categories.add(cat_normalized)
                            category_subcats[cat_normalized] = set()

                        current_category = cat_normalized

                        # Check category name length
                        if len(cat_normalized) > MAX_CATEGORY_NAME_LENGTH:
                            result.add_warning(
                                f"Category name very long ({len(cat_normalized)} chars): '{cat_normalized[:50]}...'"
                            )

                        # Check for suspicious characters
                        if re.search(r'[=|;]', cat_normalized):
                            result.add_warning(
                                f"Category '{cat_normalized}' contains suspicious characters (=, |, ;)"
                            )

                        # Check for numeric-only names
                        if cat_normalized.isdigit():
                            result.add_warning(
                                f"Category '{cat_normalized}' is numeric-only (might be accidental)"
                            )

                # Process subcategory
                if subcat_val and current_category:
                    subcat_normalized = normalize_category_name(subcat_val)

                    if subcat_normalized:
                        # Check for duplicate subcategories within same category
                        if subcat_normalized in category_subcats[current_category]:
                            result.add_error(
                                f"Duplicate subcategory '{subcat_normalized}' in category '{current_category}' "
                                f"at row {row_num + 1}"
                            )
                        else:
                            category_subcats[current_category].add(subcat_normalized)

                        # Check subcategory name length
                        if len(subcat_normalized) > MAX_CATEGORY_NAME_LENGTH:
                            result.add_warning(
                                f"Subcategory name very long: '{subcat_normalized[:50]}...'"
                            )

            # Validation checks after loading

            # Check total category count
            if len(seen_categories) > MAX_CATEGORIES:
                result.add_error(
                    f"Too many categories ({len(seen_categories)}). Maximum allowed: {MAX_CATEGORIES}"
                )

            # Check for empty categories (no subcategories)
            for cat, subcats in category_subcats.items():
                if len(subcats) == 0:
                    result.add_warning(
                        f"Category '{cat}' has no subcategories"
                    )
                elif len(subcats) == 1:
                    result.add_warning(
                        f"Category '{cat}' has only one subcategory (might be a mistake)"
                    )
                elif len(subcats) > MAX_SUBCATEGORIES_PER_CATEGORY:
                    result.add_warning(
                        f"Category '{cat}' has {len(subcats)} subcategories "
                        f"(max recommended: {MAX_SUBCATEGORIES_PER_CATEGORY})"
                    )

            # Check if template is completely empty
            if len(seen_categories) == 0:
                result.add_error(
                    "Template sheet is empty. No categories found. "
                    "Please add at least one category with subcategories."
                )

        except Exception as e:
            result.add_error(f"Error validating template: {str(e)}")
            logger.exception("Template validation failed")
        finally:
            wb.close()

        return result

    def load_category_structure_from_template(self, strict: bool = True) -> Dict[str, List[str]]:
        """
        Loads the category structure from the 'Template' sheet in the given Excel dashboard.
        Skips header rows and supports merged cells in the first column for categories.
        Now includes validation and uses robust cell reading.

        Args:
            strict: If True, raise exception on validation errors. If False, only log them.
        """
        # Validate template first
        validation = self.validate_template_structure()

        # Log any errors or warnings
        if validation.errors:
            logger.error("Template validation errors:")
            for error in validation.errors:
                logger.error(f"  - {error}")

        if validation.warnings:
            logger.warning("Template validation warnings:")
            for warning in validation.warnings:
                logger.warning(f"  - {warning}")

        # If there are critical errors and strict mode is enabled, raise exception
        if not validation.is_valid and strict:
            error_msg = (
                f"Template validation failed with {len(validation.errors)} error(s):\n"
                + "\n".join(f"  - {e}" for e in validation.errors)
            )
            raise ValueError(error_msg)

        # Load categories with improved robustness
        wb = load_workbook(self.dashboard_path, data_only=True)
        try:
            try:
                ws = wb[TEMPLATE_SHEET_NAME]
            except KeyError:
                logger.error("The worksheet 'Template' does not exist in the dashboard file.")
                available_sheets = ', '.join(wb.sheetnames) if wb.sheetnames else 'None'
                raise ValueError(
                    f"Missing 'Template' Worksheet\n"
                    f"\nThe dashboard file '{self.dashboard_path}' must contain a sheet named '{TEMPLATE_SHEET_NAME}'.\n"
                    f"\nAvailable sheets in the file: {available_sheets}\n"
                    f"\nTo fix this issue:\n"
                    f"  1. Open the dashboard file in Excel\n"
                    f"  2. Create or rename a sheet to '{TEMPLATE_SHEET_NAME}'\n"
                    f"  3. Ensure it contains your category structure (Category | Subcategory columns)\n"
                    f"  4. Save the file and try again"
                )

            categories = {}
            current_category = None

            for row in ws.iter_rows(min_row=2, max_col=2):  # skip row 1 (headers)
                cat_cell, subcat_cell = row

                # Use safe cell reading
                cat_val = safe_get_cell_value(cat_cell)
                subcat_val = safe_get_cell_value(subcat_cell)

                # Skip header rows using flexible matching
                if is_header_value(cat_val, 'category') or is_header_value(subcat_val, 'subcategory'):
                    continue

                # Detect new category with normalization
                if cat_val:
                    cat_normalized = normalize_category_name(cat_val)
                    if cat_normalized:
                        current_category = cat_normalized
                        if current_category not in categories:
                            categories[current_category] = []

                # Add subcategory with normalization
                if subcat_val and current_category:
                    subcat_normalized = normalize_category_name(subcat_val)
                    if subcat_normalized:
                        # Note: duplicates within same category are already caught by validation
                        # but we won't add them again here
                        if subcat_normalized not in categories[current_category]:
                            categories[current_category].append(subcat_normalized)

            logger.info("=== Current category structure from Template ===")
            for category, subcategories in categories.items():
                logger.info(f"{category}: {subcategories}")
            logger.info("===============================================")

            return categories
        finally:
            wb.close()



    def map_categories(self, df: pd.DataFrame) -> pd.DataFrame:
        df = df.copy()

        df['category'] = df['merchant'].map(lambda m: self.category_map.get(m, [None, None])[0])
        df['subcat']   = df['merchant'].map(lambda m: self.category_map.get(m, [None, None])[1])

        unknown = [m for m in df['merchant'].unique() if m and m not in self.category_map]
        flat_choices: List[Tuple[str, str]] = [
            (cat, sub)
            for cat, subs in self.valid_categories.items()
            for sub in subs
        ]

        try:
            for merchant in unknown:
                # Show sample rows for that merchant
                merchant_rows = df[df['merchant'] == merchant]
                sample_row = merchant_rows.iloc[0] if not merchant_rows.empty else None
                print(format_prompt(f"New merchant detected: {merchant}"))
                if sample_row is not None:
                    logger.debug(f"New merchant detected: {merchant} [date: {sample_row.get('month')}/{sample_row.get('year')}, file: {sample_row.get('source_file')}]")

                for idx, (cat, sub) in enumerate(flat_choices, start=1):
                    print(format_prompt(f"{idx}. {cat} > {sub}"))

                while True:
                    choice = input("Select category number (or 'exit'): ").strip().lower()
                    if choice == "exit":
                        print("Saving mapped categories and exiting.")
                        self.save_categories()
                        sys.exit()
                    elif not choice.isdigit():
                        print(format_prompt("Please enter a number."))
                        continue
                    idx = int(choice)
                    if 1 <= idx <= len(flat_choices):
                        cat, sub = flat_choices[idx - 1]
                        self.category_map[merchant] = [cat, sub]
                        df.loc[df['merchant'] == merchant, 'category'] = cat
                        df.loc[df['merchant'] == merchant, 'subcat'] = sub
                        break
                    else:
                        print(format_prompt("Choice out of range."))
        except KeyboardInterrupt:
            print(format_prompt("Exiting, mapped categories not saved."))
            sys.exit()
        else:
            self.save_categories()

        # Revalidate existing mappings
        df = self._handle_removed_subcategories(df)

        logger.info("Category mapping complete.")
        return df


    def _handle_removed_subcategories(self, df: pd.DataFrame) -> pd.DataFrame:
        valid_subcats = {
            (cat, sub)
            for cat, subs in self.valid_categories.items()
            for sub in subs
        }

        used_pairs = {
            (cat, sub)
            for cat, sub in zip(df['category'], df['subcat'])
            if pd.notna(cat) and pd.notna(sub)
        }

        removed_pairs = used_pairs - valid_subcats
        if not removed_pairs:
            return df

        print(format_prompt("Some previously used subcategories are no longer in the template."))
        flat_choices = [
            (cat, sub)
            for cat, subs in self.valid_categories.items()
            for sub in subs
        ]
        for idx, (cat, sub) in enumerate(flat_choices, start=1):
            print(format_prompt(f"{idx}. {cat} > {sub}"))

        for old_cat, old_sub in removed_pairs:
            print(format_prompt(f"\nSubcategory no longer exists: {old_cat} > {old_sub}"))
            while True:
                choice = input("Choose a new category number for this data (or type 'exit'): ").strip().lower()
                if choice == 'exit':
                    print("Exiting.")
                    sys.exit()
                if not choice.isdigit():
                    print(format_prompt("Please enter a number."))
                    continue
                idx = int(choice)
                if 1 <= idx <= len(flat_choices):
                    new_cat, new_sub = flat_choices[idx - 1]
                    mask = (df['category'] == old_cat) & (df['subcat'] == old_sub)
                    df.loc[mask, 'category'] = new_cat
                    df.loc[mask, 'subcat'] = new_sub
                    logger.info(f"Reassigned {mask.sum()} records from {old_cat} > {old_sub} to {new_cat} > {new_sub}")
                    break
                else:
                    print(format_prompt("Choice out of range."))

        return df

    def save_categories(self) -> None:
        with open(self.categories_path, 'w', encoding='utf-8') as f:
            json.dump(self.category_map, f, ensure_ascii=False, indent=2)
