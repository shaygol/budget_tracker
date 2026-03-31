import json
import sys
import logging
import re
import pkgutil
from typing import List, Tuple, Dict, Optional
from dataclasses import dataclass, field
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.cell import Cell
from src.previewer import format_prompt
from src.config import TEMPLATE_SHEET_NAME, MAX_CATEGORIES

logger = logging.getLogger(__name__)

# Template validation constants
MAX_SUBCATEGORIES_PER_CATEGORY = 100
MAX_CATEGORY_NAME_LENGTH = 100
HEADER_PATTERNS = {
    'category': ['נושא', 'נושא הוצאה', 'category', 'קטגוריה'],
    'subcategory': ['פירוט', 'פירוט הוצאות', 'subcategory', 'תת-קטגוריה'],
}


@dataclass
class ValidationResult:
    """Results from template validation."""
    is_valid: bool = True
    errors: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)

    def add_error(self, message: str) -> None:
        """Add an error message."""
        self.errors.append(message)
        self.is_valid = False

    def add_warning(self, message: str) -> None:
        """Add a warning message."""
        self.warnings.append(message)


def safe_get_cell_value(cell: Cell) -> Optional[str]:
    """
    Safely extract cell value, handling errors and formulas.

    Args:
        cell: Excel cell to read

    Returns:
        Cleaned string value or None if cell is empty/invalid
    """
    val = cell.value

    # Handle None
    if val is None:
        return None

    # Handle Excel errors (e.g., #REF!, #VALUE!)
    if isinstance(val, str) and val.startswith('#'):
        logger.warning(f"Excel error in cell {cell.coordinate}: {val}")
        return None

    # Convert to string and strip whitespace
    val_str = str(val).strip()

    # Return None for empty strings
    if not val_str:
        return None

    return val_str


def normalize_category_name(name: Optional[str]) -> Optional[str]:
    """
    Normalize category name for consistent matching.

    Args:
        name: Category or subcategory name

    Returns:
        Normalized name or None if invalid
    """
    if not name:
        return None

    # Strip whitespace
    name = str(name).strip()

    # Return None for empty
    if not name:
        return None

    # Collapse multiple spaces
    name = ' '.join(name.split())

    return name


def is_header_value(value: Optional[str], header_type: str) -> bool:
    """
    Check if a value looks like a header.

    Args:
        value: Cell value to check
        header_type: Type of header ('category' or 'subcategory')

    Returns:
        True if value matches header patterns
    """
    if not value:
        return False

    value_lower = value.lower().strip()
    patterns = HEADER_PATTERNS.get(header_type, [])

    return any(pattern.lower() in value_lower for pattern in patterns)


class CategoryManager:
    # Class-level cache for template structure
    _template_cache: Optional[Dict[str, List[str]]] = None
    _cache_timestamp: Optional[float] = None
    _cache_dashboard_path: Optional[Path] = None

    def __init__(self, categories_path: str | Path, dashboard_path: str | Path, strict_validation: bool = True) -> None:
        self.categories_path = categories_path
        self.dashboard_path = dashboard_path
        self.strict_validation = strict_validation
        # Merchants explicitly confirmed by user (loaded from categories.json or set in GUI flow)
        self._explicit_user_merchants = set()

        self.category_map = self._load_category_map()
        self.valid_categories = self.load_category_structure_from_template(strict=strict_validation, use_cache=True)

    def _load_category_map(self) -> Dict[str, List[str]]:
        """
        Load and merge category mappings from default and user files.
        User mappings override default mappings.
        
        Returns:
            Dictionary mapping merchant names to [category, subcategory] lists
        """
        # Start with default categories
        default_map = self._load_default_categories()
        
        # Load user categories (overrides defaults)
        user_map = self._load_user_categories()
        self._explicit_user_merchants = set(user_map.keys())
        
        # Merge: user overrides default
        merged_map = {**default_map, **user_map}
        
        logger.info(f"Loaded {len(default_map)} default categories and {len(user_map)} user categories")
        return merged_map

    def _load_default_categories(self) -> Dict[str, List[str]]:
        """
        Load default category mappings from bundled JSON resource.
        
        Returns:
            Dictionary of default mappings
        """
        try:
            # Try loading as package resource (works in compiled exe)
            data = pkgutil.get_data('src', 'default_categories.json')
            if data:
                default_map = json.loads(data.decode('utf-8'))
                logger.debug(f"Loaded {len(default_map)} default categories from package resource")
                # Ensure all entries are properly formatted [category, subcategory]
                return {
                    merchant: mapping if isinstance(mapping, list) and len(mapping) >= 2 else []
                    for merchant, mapping in default_map.items()
                }
        except Exception as e:
            logger.debug(f"Could not load from package resource: {e}")
        
        # Fallback: try loading from file system (development/testing)
        try:
            # Try multiple possible locations
            possible_paths = [
                Path(__file__).parent / 'default_categories.json',  # Same dir as this file
                Path(__file__).parent.parent / 'src' / 'default_categories.json',  # From project root
            ]
            
            for default_file in possible_paths:
                if default_file.exists():
                    with open(default_file, 'r', encoding='utf-8') as f:
                        default_map = json.load(f)
                        logger.debug(f"Loaded {len(default_map)} default categories from {default_file}")
                        return {
                            merchant: mapping if isinstance(mapping, list) and len(mapping) >= 2 else []
                            for merchant, mapping in default_map.items()
                        }
        except Exception as e2:
            logger.warning(f"Could not load default categories from file: {e2}")
        
        logger.warning("No default categories loaded")
        return {}

    def _load_user_categories(self) -> Dict[str, List[str]]:
        """
        Load user-specific category mappings from UserFiles/categories.json
        
        Returns:
            Dictionary of user mappings
        """
        try:
            with open(self.categories_path, 'r', encoding='utf-8') as f:
                user_map = json.load(f)
                # Clean up: remove _default flag if present and ensure proper format
                cleaned_map = {}
                for merchant, mapping in user_map.items():
                    if isinstance(mapping, list) and len(mapping) >= 2:
                        # Keep only [category, subcategory], strip _default flag
                        cleaned_map[merchant] = mapping[:2]
                return cleaned_map
        except (FileNotFoundError, json.JSONDecodeError):
            logger.warning(f"User categories file not found or invalid: {self.categories_path}. Starting with empty map.")
            return {}

    def validate_template_structure(self) -> ValidationResult:
        """
        Validate the template structure for data quality issues.

        Returns:
            ValidationResult with errors and warnings
        """
        result = ValidationResult()

        # Check if file exists
        if not Path(self.dashboard_path).exists():
            result.add_error(f"Dashboard file not found: {self.dashboard_path}")
            return result

        wb = load_workbook(self.dashboard_path, data_only=True)
        try:
            # Check if Template sheet exists
            if TEMPLATE_SHEET_NAME not in wb.sheetnames:
                result.add_error(
                    f"Template sheet '{TEMPLATE_SHEET_NAME}' not found. "
                    f"Available sheets: {', '.join(wb.sheetnames)}"
                )
                return result

            ws = wb[TEMPLATE_SHEET_NAME]

            # Track for duplicate detection
            seen_categories = set()
            category_subcats = {}  # Track subcats per category for duplicate detection

            current_category = None
            row_num = 0

            for row in ws.iter_rows(min_row=2, max_col=2):
                row_num += 1
                cat_cell, subcat_cell = row

                # Safe cell reading
                cat_val = safe_get_cell_value(cat_cell)
                subcat_val = safe_get_cell_value(subcat_cell)

                # Skip header rows
                if is_header_value(cat_val, 'category') or is_header_value(subcat_val, 'subcategory'):
                    continue

                # Process category
                if cat_val:
                    cat_normalized = normalize_category_name(cat_val)

                    if cat_normalized:
                        # Check for duplicate categories
                        if cat_normalized in seen_categories:
                            result.add_error(
                                f"Duplicate category '{cat_normalized}' found at row {row_num + 1}"
                            )
                        else:
                            seen_categories.add(cat_normalized)
                            category_subcats[cat_normalized] = set()

                        current_category = cat_normalized

                        # Check category name length
                        if len(cat_normalized) > MAX_CATEGORY_NAME_LENGTH:
                            result.add_warning(
                                f"Category name very long ({len(cat_normalized)} chars): '{cat_normalized[:50]}...'"
                            )

                        # Check for suspicious characters
                        if re.search(r'[=|;]', cat_normalized):
                            result.add_warning(
                                f"Category '{cat_normalized}' contains suspicious characters (=, |, ;)"
                            )

                        # Check for numeric-only names
                        if cat_normalized.isdigit():
                            result.add_warning(
                                f"Category '{cat_normalized}' is numeric-only (might be accidental)"
                            )

                # Process subcategory
                if subcat_val and current_category:
                    subcat_normalized = normalize_category_name(subcat_val)

                    if subcat_normalized:
                        # Check for duplicate subcategories within same category
                        if subcat_normalized in category_subcats[current_category]:
                            result.add_error(
                                f"Duplicate subcategory '{subcat_normalized}' in category '{current_category}' "
                                f"at row {row_num + 1}"
                            )
                        else:
                            category_subcats[current_category].add(subcat_normalized)

                        # Check subcategory name length
                        if len(subcat_normalized) > MAX_CATEGORY_NAME_LENGTH:
                            result.add_warning(
                                f"Subcategory name very long: '{subcat_normalized[:50]}...'"
                            )

            # Validation checks after loading

            # Check total category count
            if len(seen_categories) > MAX_CATEGORIES:
                result.add_error(
                    f"Too many categories ({len(seen_categories)}). Maximum allowed: {MAX_CATEGORIES}"
                )

            # Check for empty categories (no subcategories)
            for cat, subcats in category_subcats.items():
                if len(subcats) == 0:
                    result.add_warning(
                        f"Category '{cat}' has no subcategories"
                    )
                elif len(subcats) == 1:
                    result.add_warning(
                        f"Category '{cat}' has only one subcategory (might be a mistake)"
                    )
                elif len(subcats) > MAX_SUBCATEGORIES_PER_CATEGORY:
                    result.add_warning(
                        f"Category '{cat}' has {len(subcats)} subcategories "
                        f"(max recommended: {MAX_SUBCATEGORIES_PER_CATEGORY})"
                    )

            # Check if template is completely empty
            if len(seen_categories) == 0:
                result.add_error(
                    "Template sheet is empty. No categories found. "
                    "Please add at least one category with subcategories."
                )

        except Exception as e:
            result.add_error(f"Error validating template: {str(e)}")
            logger.exception("Template validation failed")
        finally:
            wb.close()

        return result

    def load_category_structure_from_template(self, strict: bool = True, use_cache: bool = True) -> Dict[str, List[str]]:
        """
        Loads the category structure from the 'Template' sheet in the given Excel dashboard.
        Skips header rows and supports merged cells in the first column for categories.
        Now includes validation and uses robust cell reading.
        Uses in-memory cache to reduce file I/O.

        Args:
            strict: If True, raise exception on validation errors. If False, only log them.
            use_cache: If True, use cached template structure if available and valid.

        Returns:
            Dictionary mapping category names to lists of subcategory names
        """
        dashboard_path = Path(self.dashboard_path)

        # Check cache validity
        if use_cache and self._template_cache is not None:
            if self._cache_dashboard_path == dashboard_path:
                # Check if file has been modified since cache was created
                try:
                    current_mtime = dashboard_path.stat().st_mtime
                    if self._cache_timestamp is not None and current_mtime <= self._cache_timestamp:
                        logger.debug("Using cached template structure")
                        return self._template_cache.copy()  # Return a copy to prevent mutation
                except (OSError, FileNotFoundError):
                    # File doesn't exist or can't be accessed, invalidate cache
                    self._invalidate_cache()

        # Cache miss or invalid - load from file
        # Validate template first
        validation = self.validate_template_structure()

        # Log any errors or warnings
        if validation.errors:
            logger.error("Template validation errors:")
            for error in validation.errors:
                logger.error(f"  - {error}")

        if validation.warnings:
            logger.warning("Template validation warnings:")
            for warning in validation.warnings:
                logger.warning(f"  - {warning}")

        # If there are critical errors and strict mode is enabled, raise exception
        if not validation.is_valid and strict:
            error_msg = (
                f"Template validation failed with {len(validation.errors)} error(s):\n"
                + "\n".join(f"  - {e}" for e in validation.errors)
            )
            raise ValueError(error_msg)

        # Load categories with improved robustness
        wb = load_workbook(self.dashboard_path, data_only=True)
        try:
            try:
                ws = wb[TEMPLATE_SHEET_NAME]
            except KeyError:
                logger.error("The worksheet 'Template' does not exist in the dashboard file.")
                available_sheets = ', '.join(wb.sheetnames) if wb.sheetnames else 'None'
                raise ValueError(
                    f"Missing 'Template' Worksheet\n"
                    f"\nThe dashboard file '{self.dashboard_path}' must contain a sheet named '{TEMPLATE_SHEET_NAME}'.\n"
                    f"\nAvailable sheets in the file: {available_sheets}\n"
                    f"\nTo fix this issue:\n"
                    f"  1. Open the dashboard file in Excel\n"
                    f"  2. Create or rename a sheet to '{TEMPLATE_SHEET_NAME}'\n"
                    f"  3. Ensure it contains your category structure (Category | Subcategory columns)\n"
                    f"  4. Save the file and try again"
                )

            categories = {}
            current_category = None

            for row in ws.iter_rows(min_row=2, max_col=2):  # skip row 1 (headers)
                cat_cell, subcat_cell = row

                # Use safe cell reading
                cat_val = safe_get_cell_value(cat_cell)
                subcat_val = safe_get_cell_value(subcat_cell)

                # Skip header rows using flexible matching
                if is_header_value(cat_val, 'category') or is_header_value(subcat_val, 'subcategory'):
                    continue

                # Detect new category with normalization
                if cat_val:
                    cat_normalized = normalize_category_name(cat_val)
                    if cat_normalized:
                        current_category = cat_normalized
                        if current_category not in categories:
                            categories[current_category] = []

                # Add subcategory with normalization
                if subcat_val and current_category:
                    subcat_normalized = normalize_category_name(subcat_val)
                    if subcat_normalized:
                        # Note: duplicates within same category are already caught by validation
                        # but we won't add them again here
                        if subcat_normalized not in categories[current_category]:
                            categories[current_category].append(subcat_normalized)

            logger.info("=== Current category structure from Template ===")
            for category, subcategories in categories.items():
                logger.info(f"{category}: {subcategories}")
            logger.info("===============================================")

            # Update cache
            if use_cache:
                try:
                    self._template_cache = categories.copy()
                    self._cache_timestamp = dashboard_path.stat().st_mtime
                    self._cache_dashboard_path = dashboard_path
                    logger.debug("Template structure cached")
                except (OSError, FileNotFoundError):
                    # If we can't get mtime, don't cache
                    pass

            return categories
        finally:
            wb.close()

    @classmethod
    def _invalidate_cache(cls) -> None:
        """
        Invalidate the template structure cache.

        Call this when the dashboard file is known to have been modified.
        """
        cls._template_cache = None
        cls._cache_timestamp = None
        cls._cache_dashboard_path = None
        logger.debug("Template cache invalidated")

    @classmethod
    def get_cached_categories(cls) -> Optional[Dict[str, List[str]]]:
        """
        Get cached category structure without loading from file.

        Returns:
            Cached category structure or None if cache is invalid/empty
        """
        if cls._template_cache is not None:
            return cls._template_cache.copy()
        return None



    def map_categories(self, df: pd.DataFrame) -> pd.DataFrame:
        df = df.copy()

        df['category'] = df['merchant'].map(lambda m: self.category_map.get(m, [None, None])[0])
        df['subcat']   = df['merchant'].map(lambda m: self.category_map.get(m, [None, None])[1])

        unknown = [m for m in df['merchant'].unique() if m and m not in self.category_map]
        flat_choices: List[Tuple[str, str]] = [
            (cat, sub)
            for cat, subs in self.valid_categories.items()
            for sub in subs
        ]

        try:
            for merchant in unknown:
                # Show sample rows for that merchant
                merchant_rows = df[df['merchant'] == merchant]
                sample_row = merchant_rows.iloc[0] if not merchant_rows.empty else None
                print(format_prompt(f"New merchant detected: {merchant}"))
                if sample_row is not None:
                    logger.debug(f"New merchant detected: {merchant} [date: {sample_row.get('month')}/{sample_row.get('year')}, file: {sample_row.get('source_file')}]")

                for idx, (cat, sub) in enumerate(flat_choices, start=1):
                    print(format_prompt(f"{idx}. {cat} > {sub}"))

                while True:
                    choice = input("Select category number (or 'exit'): ").strip().lower()
                    if choice == "exit":
                        print("Saving mapped categories and exiting.")
                        self.save_categories()
                        sys.exit()
                    elif not choice.isdigit():
                        print(format_prompt("Please enter a number."))
                        continue
                    idx = int(choice)
                    if 1 <= idx <= len(flat_choices):
                        cat, sub = flat_choices[idx - 1]
                        self.category_map[merchant] = [cat, sub]
                        self.mark_user_confirmed(merchant)
                        df.loc[df['merchant'] == merchant, 'category'] = cat
                        df.loc[df['merchant'] == merchant, 'subcat'] = sub
                        break
                    else:
                        print(format_prompt("Choice out of range."))
        except KeyboardInterrupt:
            print(format_prompt("Exiting, mapped categories not saved."))
            sys.exit()
        else:
            self.save_categories()

        # Revalidate existing mappings
        df = self._handle_removed_subcategories(df)

        logger.info("Category mapping complete.")
        return df


    def _handle_removed_subcategories(self, df: pd.DataFrame) -> pd.DataFrame:
        valid_subcats = {
            (cat, sub)
            for cat, subs in self.valid_categories.items()
            for sub in subs
        }

        used_pairs = {
            (cat, sub)
            for cat, sub in zip(df['category'], df['subcat'])
            if pd.notna(cat) and pd.notna(sub)
        }

        removed_pairs = used_pairs - valid_subcats
        if not removed_pairs:
            return df

        print(format_prompt("Some previously used subcategories are no longer in the template."))
        flat_choices = [
            (cat, sub)
            for cat, subs in self.valid_categories.items()
            for sub in subs
        ]
        for idx, (cat, sub) in enumerate(flat_choices, start=1):
            print(format_prompt(f"{idx}. {cat} > {sub}"))

        for old_cat, old_sub in removed_pairs:
            print(format_prompt(f"\nSubcategory no longer exists: {old_cat} > {old_sub}"))
            while True:
                choice = input("Choose a new category number for this data (or type 'exit'): ").strip().lower()
                if choice == 'exit':
                    print("Exiting.")
                    sys.exit()
                if not choice.isdigit():
                    print(format_prompt("Please enter a number."))
                    continue
                idx = int(choice)
                if 1 <= idx <= len(flat_choices):
                    new_cat, new_sub = flat_choices[idx - 1]
                    mask = (df['category'] == old_cat) & (df['subcat'] == old_sub)
                    df.loc[mask, 'category'] = new_cat
                    df.loc[mask, 'subcat'] = new_sub
                    logger.info(f"Reassigned {mask.sum()} records from {old_cat} > {old_sub} to {new_cat} > {new_sub}")
                    break
                else:
                    print(format_prompt("Choice out of range."))

        return df

    def find_similar_merchant(self, merchant_name: str) -> Optional[Tuple[str, str]]:
        """
        Find a similar merchant in category_map and return its category.

        Uses multiple matching strategies:
        1. Exact match
        2. Prefix matching (e.g., "AMAZON.COM" matches "AMAZON PRIME")
        3. Keyword matching (e.g., "ביטוח חובה" matches "ביטוח")

        Args:
            merchant_name: Name of the merchant to find a match for

        Returns:
            Tuple of (category, subcategory) if match found, None otherwise
        """
        if not merchant_name:
            return None

        merchant_upper = merchant_name.upper().strip()

        # PRIORITY 1: Try exact match
        if self.category_map:
            for existing_merchant, mapping in self.category_map.items():
                # Skip invalid entries
                if not existing_merchant or not isinstance(mapping, (list, tuple)) or len(mapping) < 2:
                    continue
                
                cat, sub = mapping[0], mapping[1]
                
                if existing_merchant.upper().strip() == merchant_upper:
                    logger.debug(f"Found exact match for '{merchant_name}': {cat} > {sub}")
                    return (cat, sub)

        # PRIORITY 2 & 3: Try both prefix and substring matching, pick the best
        prefix_match = None
        prefix_length = 0
        prefix_merchant = None
        
        substring_match = None
        substring_length = 0
        substring_merchant = None

        if self.category_map:
            for existing_merchant, mapping in self.category_map.items():
                # Skip invalid entries
                if not existing_merchant or not isinstance(mapping, (list, tuple)) or len(mapping) < 2:
                    continue
                    
                cat, sub = mapping[0], mapping[1]
                existing_upper = existing_merchant.upper().strip()

                # Check prefix matching (minimum 4 characters)
                if len(merchant_upper) >= 4 and len(existing_upper) >= 4:
                    if merchant_upper.startswith(existing_upper[:4]) or existing_upper.startswith(merchant_upper[:4]):
                        # Calculate common prefix length
                        common_length = 0
                        for i in range(min(len(merchant_upper), len(existing_upper))):
                            if merchant_upper[i] == existing_upper[i]:
                                common_length += 1
                            else:
                                break

                        if common_length >= 4 and common_length > prefix_length:
                            prefix_match = (cat, sub)
                            prefix_length = common_length
                            prefix_merchant = existing_merchant
                
                # Check longest common substring
                max_len = 0
                merchant_len = len(merchant_upper)
                existing_len = len(existing_upper)
                
                # Use dynamic programming to find longest common substring
                for i in range(merchant_len):
                    for j in range(existing_len):
                        length = 0
                        while (i + length < merchant_len and 
                               j + length < existing_len and 
                               merchant_upper[i + length] == existing_upper[j + length]):
                            length += 1
                        if length > max_len:
                            max_len = length
                
                # Require minimum length of 4 characters for match
                if max_len >= 4 and max_len > substring_length:
                    substring_match = (cat, sub)
                    substring_length = max_len
                    substring_merchant = existing_merchant

        # Choose the best match: prefer longer match
        if substring_length > prefix_length and substring_match:
            logger.debug(f"Found substring match for '{merchant_name}' with '{substring_merchant}' (length {substring_length}): {substring_match[0]} > {substring_match[1]}")
            return substring_match
        elif prefix_match:
            logger.debug(f"Found prefix match for '{merchant_name}' with '{prefix_merchant}' (length {prefix_length}): {prefix_match[0]} > {prefix_match[1]}")
            return prefix_match

        return None

    def save_categories(self) -> None:
        """
        Save only user-confirmed category mappings to file.
        Default categories are not saved by default, except when explicitly confirmed by user.
        Saves mappings that are:
        1. Explicitly confirmed by user, OR
        2. Different from defaults, OR
        3. Not present in defaults.
        """
        # Load defaults for comparison
        default_map = self._load_default_categories()
        
        # Save only mappings that are different from defaults or not in defaults
        user_only_map = {}
        for merchant, mapping in self.category_map.items():
            if isinstance(mapping, list) and len(mapping) >= 2:
                # Keep only [category, subcategory]
                clean_mapping = mapping[:2]
                
                if (
                    merchant in self._explicit_user_merchants
                    or merchant not in default_map
                    or clean_mapping != default_map.get(merchant, [])[:2]
                ):
                    user_only_map[merchant] = clean_mapping
        
        with open(self.categories_path, 'w', encoding='utf-8') as f:
            json.dump(user_only_map, f, ensure_ascii=False, indent=2)
        
        logger.info(f"Saved {len(user_only_map)} user-specific category mappings")

    def mark_user_confirmed(self, merchant: str) -> None:
        """Mark merchant mapping as explicitly confirmed by user."""
        if merchant:
            self._explicit_user_merchants.add(merchant)
