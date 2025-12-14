"""
File operation utilities for error handling and validation.
"""
import os
import logging
from pathlib import Path
from typing import Optional
import platform

logger = logging.getLogger(__name__)


def is_file_locked(file_path: Path) -> bool:
    """
    Check if a file is locked/in use (Windows-specific).

    On Windows, attempts to open the file in exclusive mode. If successful,
    the file is not locked. On other platforms, checks if file is readable.

    Args:
        file_path: Path to the file to check

    Returns:
        True if file appears to be locked, False otherwise
    """
    if not file_path.exists():
        return False

    try:
        if platform.system() == 'Windows':
            # On Windows, try to open file in exclusive mode
            try:
                # Try to open in append mode - if file is locked, this will fail
                with open(file_path, 'a'):
                    pass
                return False
            except (IOError, OSError, PermissionError):
                return True
        else:
            # On Unix-like systems, check if file is readable
            return not os.access(file_path, os.R_OK)
    except Exception as e:
        logger.warning(f"Error checking file lock status for {file_path}: {e}")
        # If we can't determine, assume not locked to avoid blocking operations
        return False


def validate_dashboard_integrity(file_path: Path) -> tuple[bool, Optional[str]]:
    """
    Validate dashboard file structure and integrity.

    Checks if the file exists, can be opened, and contains a Template sheet.

    Args:
        file_path: Path to dashboard Excel file

    Returns:
        Tuple of (is_valid, error_message). is_valid is True if file is valid,
        error_message is None if valid, otherwise contains error description.
    """
    if not file_path.exists():
        return False, f"Dashboard file not found: {file_path}"

    if not file_path.is_file():
        return False, f"Path is not a file: {file_path}"

    # Check if file is locked
    if is_file_locked(file_path):
        return False, f"Dashboard file is locked (may be open in Excel): {file_path}"

    # Try to open and validate structure
    try:
        from openpyxl import load_workbook
        from src.config import TEMPLATE_SHEET_NAME

        wb = load_workbook(file_path, read_only=True, data_only=True)
        try:
            if TEMPLATE_SHEET_NAME not in wb.sheetnames:
                return False, f"Template sheet '{TEMPLATE_SHEET_NAME}' not found in dashboard"

            # Check if template sheet has data
            ws = wb[TEMPLATE_SHEET_NAME]
            if ws.max_row < 2:
                return False, "Template sheet appears to be empty"

            return True, None
        finally:
            wb.close()
    except Exception as e:
        return False, f"Error reading dashboard file: {str(e)}"


def get_user_friendly_error(exception: Exception) -> str:
    """
    Convert technical exceptions to user-friendly error messages.

    Args:
        exception: Exception object to convert

    Returns:
        User-friendly error message string
    """
    error_type = type(exception).__name__
    error_msg = str(exception)

    # File not found errors
    if isinstance(exception, FileNotFoundError):
        return f"File not found. Please check that the file exists and the path is correct.\n\nDetails: {error_msg}"

    # Permission errors
    if isinstance(exception, PermissionError):
        return f"Permission denied. The file may be open in another program (like Excel).\n\nPlease close the file and try again.\n\nDetails: {error_msg}"

    # IO errors
    if isinstance(exception, IOError) or isinstance(exception, OSError):
        if "locked" in error_msg.lower() or "being used" in error_msg.lower():
            return f"File is locked. The file may be open in Excel or another program.\n\nPlease close the file and try again."
        return f"File operation failed. Please check that the file is not corrupted and you have permission to access it.\n\nDetails: {error_msg}"

    # Excel-specific errors
    if "openpyxl" in error_type or "xlrd" in error_type or "Excel" in error_type:
        return f"Excel file error. The file may be corrupted or in an unsupported format.\n\nPlease verify the file is a valid Excel file (.xlsx or .xls).\n\nDetails: {error_msg}"

    # JSON errors
    if isinstance(exception, ValueError) and "json" in error_msg.lower():
        return f"Invalid data format. The categories file may be corrupted.\n\nPlease check the file or delete it to start fresh.\n\nDetails: {error_msg}"

    # Generic error
    return f"An error occurred: {error_msg}\n\nIf this problem persists, please check the log file for more details."


def check_file_permissions(file_path: Path, operation: str = 'read') -> tuple[bool, Optional[str]]:
    """
    Check if file permissions allow the specified operation.

    Args:
        file_path: Path to the file
        operation: Operation to check ('read', 'write', 'execute')

    Returns:
        Tuple of (has_permission, error_message)
    """
    if not file_path.exists():
        return False, f"File does not exist: {file_path}"

    try:
        if operation == 'read':
            if not os.access(file_path, os.R_OK):
                return False, f"No read permission for file: {file_path}"
        elif operation == 'write':
            if not os.access(file_path, os.W_OK):
                return False, f"No write permission for file: {file_path}"
        elif operation == 'execute':
            if not os.access(file_path, os.X_OK):
                return False, f"No execute permission for file: {file_path}"

        return True, None
    except Exception as e:
        return False, f"Error checking permissions: {str(e)}"

