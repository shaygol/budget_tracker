import pandas as pd
import os
import logging
from pathlib import Path
from typing import Optional, Dict, Tuple, Callable, List
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet
from src.config import TEMPLATE_SHEET_NAME, DASHBOARD_BACKUP_DIR

logger = logging.getLogger(__name__)
SUMMARY_LABELS = {"Summary", "סיכום"}


class DashboardWriter:
    def __init__(self, dashboard_path: Path | str) -> None:
        self.dashboard_path = dashboard_path
        self.template_sheet_name = TEMPLATE_SHEET_NAME
        self.user_decisions = {}

    def update(self, summary_df: pd.DataFrame, conflict_resolver: Optional[Callable[[str], str]] = None) -> None:
        """
        Update dashboard with summary data.

        Args:
            summary_df: DataFrame with transaction summary data
            conflict_resolver: Optional callback function to resolve data conflicts

        Raises:
            TransactionValidationError: If validation fails
        """
        if not self._validate_summary(summary_df):
            return

        # Comprehensive validation
        is_valid, errors = self.validate_summary_data(summary_df)
        if not is_valid:
            error_msg = "Transaction validation failed:\n" + "\n".join(f"  - {e}" for e in errors[:10])
            logger.error(error_msg)
            from src.transaction_validator import TransactionValidationError
            raise TransactionValidationError(error_msg)

        self.conflict_resolver = conflict_resolver

        years = sorted(summary_df['year'].unique())
        file_exists = os.path.exists(self.dashboard_path)

        if not file_exists:
            logger.error(f"Dashboard file not found: {self.dashboard_path}")
            raise FileNotFoundError(f"Dashboard file not found: {self.dashboard_path}")

        # Check if file is locked
        from src.file_utils import is_file_locked
        if is_file_locked(Path(self.dashboard_path)):
            error_msg = f"Dashboard file is locked (may be open in Excel): {self.dashboard_path}"
            logger.error(error_msg)
            raise PermissionError(error_msg)

        # Create backup before modifying dashboard
        import shutil
        from datetime import datetime

        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            # Ensure backup directory exists
            os.makedirs(DASHBOARD_BACKUP_DIR, exist_ok=True)

            stem = Path(self.dashboard_path).stem
            suffix = Path(self.dashboard_path).suffix
            backup_filename = f"{stem}_{timestamp}{suffix}"
            backup_path = DASHBOARD_BACKUP_DIR / backup_filename

            shutil.copy2(self.dashboard_path, backup_path)
            logger.info(f"Created backup: {backup_path}")
        except (IOError, OSError, PermissionError) as e:
            logger.warning(f"Failed to create backup: {e}")
            # Continue even if backup fails, but log the warning
        except Exception as e:
            logger.warning(f"Unexpected error creating backup: {e}")

        try:
            wb = load_workbook(self.dashboard_path)
        except PermissionError as e:
            logger.error(f"Permission denied opening dashboard: {e}")
            raise
        except (IOError, OSError) as e:
            logger.error(f"IO error opening dashboard: {e}")
            raise
        except Exception as e:
            logger.error(f"Error opening dashboard file: {e}")
            raise

        try:
            if self.template_sheet_name not in wb.sheetnames:
                logger.error(f"Template sheet '{self.template_sheet_name}' not found in dashboard.")
                return

            for year in years:
                sheet_name = str(year)
                if sheet_name not in wb.sheetnames:
                    self._clone_template_sheet(wb, sheet_name)

                ws = wb[sheet_name]
                self._populate_sheet(ws, year, summary_df[summary_df['year'] == year])

            wb.save(self.dashboard_path)
        finally:
            wb.close()

    def _validate_summary(self, df: pd.DataFrame) -> bool:
        required_cols = {'year', 'month', 'category', 'subcat', 'monthly_amount'}
        if not required_cols.issubset(df.columns):
            logger.error(f"Missing required columns: {required_cols - set(df.columns)}")
            return False
        if df.empty:
            logger.warning("Summary DataFrame is empty. Nothing will be written to the dashboard.")
            return False
        return True

    def validate_summary_data(self, summary_df: pd.DataFrame) -> Tuple[bool, List[str]]:
        """
        Validate summary DataFrame before writing to dashboard.

        Args:
            summary_df: DataFrame with transaction summary

        Returns:
            Tuple of (is_valid, list_of_errors)
        """
        from datetime import datetime
        errors = []

        # Basic validation (already done by _validate_summary, but return errors for consistency)
        if not self._validate_summary(summary_df):
            errors.append("Summary validation failed - missing columns or empty DataFrame")
            return False, errors

        # Validate data types
        if not pd.api.types.is_numeric_dtype(summary_df['year']):
            errors.append("Year column must be numeric")
        if not pd.api.types.is_numeric_dtype(summary_df['month']):
            errors.append("Month column must be numeric")
        if not pd.api.types.is_numeric_dtype(summary_df['monthly_amount']):
            errors.append("Monthly amount column must be numeric")

        # Validate year range
        current_year = datetime.now().year
        if summary_df['year'].min() < 2000 or summary_df['year'].max() > current_year + 1:
            errors.append(f"Year values must be between 2000 and {current_year + 1}")

        # Validate month range
        invalid_months = summary_df[(summary_df['month'] < 1) | (summary_df['month'] > 12)]
        if not invalid_months.empty:
            errors.append(f"Invalid month values found: {invalid_months['month'].unique().tolist()}")

        # Warn about negative amounts (e.g. refunds) but don't fail validation
        negative_amounts = summary_df[summary_df['monthly_amount'] < 0]
        if not negative_amounts.empty:
            logger.warning(f"Found {len(negative_amounts)} transaction(s) with negative amounts (possibly refunds)")

        is_valid = len(errors) == 0
        return is_valid, errors

    def _clone_template_sheet(self, wb: Workbook, new_sheet_name: str) -> None:
        source = wb[self.template_sheet_name]
        target = wb.copy_worksheet(source)
        target.title = new_sheet_name
        target.sheet_view.rightToLeft = source.sheet_view.rightToLeft
        logger.info(f"Created new sheet by cloning template: {new_sheet_name}")

    def _populate_sheet(self, ws: Worksheet, year: int, df: pd.DataFrame) -> None:
        # Map month numbers to column indexes (e.g., Jan -> col 3)
        month_columns = {str(m): 2 + (m - 1) for m in range(1, 13)}

        # Build mapping of existing subcategories to rows
        category_ranges = self._get_category_row_ranges(ws)
        existing_map = self._build_subcat_location_map(ws, category_ranges)

        for _, row in df.iterrows():
            cat = row['category']
            subcat = row['subcat']
            month = str(int(row['month']))
            amount = row['monthly_amount']
            col = month_columns[month]

            if cat not in category_ranges:
                logger.info(f"New category detected: {cat}. Adding it.")
                self._add_new_category(ws, cat)
                category_ranges = self._get_category_row_ranges(ws)

            if (cat, subcat) not in existing_map:
                logger.info(f"New subcategory '{subcat}' under '{cat}' detected. Adding it.")
                self._add_new_subcategory(ws, cat, subcat, existing_map, category_ranges)
                existing_map = self._build_subcat_location_map(ws, category_ranges)

            row_idx = existing_map[(cat, subcat)]
            cell = ws.cell(row=row_idx, column=col + 1)
            existing_value = cell.value

            if existing_value is not None and existing_value != 0 and existing_value != "":
                month_key = f"{year}-{month}"
                if month_key not in self.user_decisions:
                    decision = self._prompt_user_decision(month_key, self.conflict_resolver)
                    self.user_decisions[month_key] = decision
                else:
                    decision = self.user_decisions[month_key]

                if decision == "override":
                    logger.debug(f"Overriding cell {cell.coordinate} with {amount}")
                    cell.value = amount
                elif decision == "add":
                    try:
                        new_val = float(existing_value) + amount
                        logger.debug(f"Adding to cell {cell.coordinate}: {existing_value} + {amount} = {new_val}")
                        cell.value = new_val
                    except Exception:
                        cell.value = amount
                elif decision == "skip":
                    logger.debug(f"Skipping cell {cell.coordinate}")
                    continue
            else:
                logger.debug(f"Writing to new cell {cell.coordinate}: {amount}")
                cell.value = amount

            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=False)

        logger.info(f"Dashboard sheet populated for year {year}")


    def _prompt_user_decision(self, month_key: str, conflict_resolver: Optional[Callable[[str], str]] = None) -> str:
        # Use callback if provided (GUI mode), otherwise use input (CLI mode)
        if conflict_resolver:
            return conflict_resolver(month_key)

        print(f"\nData already exists for {month_key}. Choose how to handle it:")
        print("1. Override existing data")
        print("2. Add to existing data")
        print("3. Skip this month")
        choice = input("Enter 1/2/3: ").strip()
        return {"1": "override", "2": "add", "3": "skip"}.get(choice, "skip")

    def _get_category_row_ranges(self, ws: Worksheet) -> Dict[str, Tuple[int, int]]:
        # Returns a map: category name -> (start_row, end_row)
        ranges = {}
        current_cat = None
        start = None
        summary_row = self._find_summary_row(ws)
        last_data_row = (summary_row - 1) if summary_row else ws.max_row

        for idx in range(2, last_data_row + 1):
            cat_value = ws.cell(row=idx, column=1).value
            if cat_value:
                if current_cat and start is not None:
                    ranges[current_cat] = (start, idx - 1)
                current_cat = cat_value
                start = idx

        if current_cat and start is not None:
            ranges[current_cat] = (start, last_data_row)
        return ranges

    def _build_subcat_location_map(self, ws: Worksheet, cat_ranges: Dict[str, Tuple[int, int]]) -> Dict[Tuple[str, str], int]:
        # Returns a map: (category, subcategory) -> row index
        mapping = {}
        for cat, (start, end) in cat_ranges.items():
            for r in range(start, end + 1):
                subcat_cell = ws.cell(row=r, column=2)
                val = subcat_cell.value
                if isinstance(val, str) and val.strip():
                    clean_subcat = val.strip()
                    mapping[(cat, clean_subcat)] = r
                    logger.debug(f"Mapped: ({cat}, {clean_subcat}) -> row {r}")
        logger.debug(f"Total subcategories mapped: {len(mapping)}")
        return mapping

    def _add_new_category(self, ws: Worksheet, category: str) -> None:
        # Add new categories before the summary row so formulas continue to include them.
        summary_row = self._find_summary_row(ws)
        insert_at = summary_row if summary_row else ws.max_row + 1
        self._insert_rows_preserving_merges(ws, insert_at)
        ws.cell(row=insert_at, column=1, value=category)
        for col in range(2, 15):
            ws.cell(row=insert_at, column=col, value="")

    def _add_new_subcategory(self, ws: Worksheet, category: str, subcat: str,
                             subcat_map: Dict[Tuple[str, str], int], cat_ranges: Dict[str, Tuple[int, int]]) -> None:
        # Insert the new subcategory below the category group and above the summary row.
        start, end = cat_ranges[category]
        insert_at = end + 1

        logger.debug(f"Inserting new subcategory '{subcat}' at row {insert_at} (Category range: {start}-{end})")
        self._insert_rows_preserving_merges(ws, insert_at)
        ws.cell(row=insert_at, column=2, value=subcat)
        for col in range(3, 15):
            ws.cell(row=insert_at, column=col, value="")
        # Re-merge the category header cell to extend one more row
        if start < end:
            ws.unmerge_cells(start_row=start, end_row=end, start_column=1, end_column=1)
        ws.merge_cells(start_row=start, end_row=end + 1, start_column=1, end_column=1)
        cat_ranges[category] = (start, end + 1)

    def _find_summary_row(self, ws: Worksheet) -> Optional[int]:
        """Find the summary row that is kept at the bottom of the dashboard."""
        for merged_range in ws.merged_cells.ranges:
            if (
                merged_range.min_col == 1
                and merged_range.max_col == 2
                and merged_range.min_row == merged_range.max_row
            ):
                value = ws.cell(row=merged_range.min_row, column=1).value
                if isinstance(value, str) and value.strip() in SUMMARY_LABELS:
                    return merged_range.min_row

        for row_idx in range(2, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=1).value
            if isinstance(value, str) and value.strip() in SUMMARY_LABELS:
                return row_idx

        return None

    def _insert_rows_preserving_merges(self, ws: Worksheet, insert_at: int, amount: int = 1) -> None:
        """Insert rows while keeping merged ranges aligned with their logical content."""
        merged_ranges = [
            (rng.min_row, rng.max_row, rng.min_col, rng.max_col)
            for rng in list(ws.merged_cells.ranges)
        ]

        for min_row, max_row, min_col, max_col in merged_ranges:
            ws.unmerge_cells(
                start_row=min_row,
                end_row=max_row,
                start_column=min_col,
                end_column=max_col,
            )

        ws.insert_rows(insert_at, amount)

        for min_row, max_row, min_col, max_col in merged_ranges:
            if max_row < insert_at:
                new_min_row, new_max_row = min_row, max_row
            elif min_row >= insert_at:
                new_min_row, new_max_row = min_row + amount, max_row + amount
            else:
                new_min_row, new_max_row = min_row, max_row + amount

            ws.merge_cells(
                start_row=new_min_row,
                end_row=new_max_row,
                start_column=min_col,
                end_column=max_col,
            )
